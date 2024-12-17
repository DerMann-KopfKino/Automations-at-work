from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import UnexpectedAlertPresentException
from selenium.common.exceptions import ElementClickInterceptedException
from pdfminer.pdfparser import PDFSyntaxError
from selenium.webdriver.support.ui import Select
from PyPDF2 import PdfWriter, PdfReader, PdfFileMerger
import os
import glob
import openpyxl
import time
import datetime
import numpy as np
import pandas as pd
import pdfkit
import pyautogui
import pdfplumber
import re
import sqlite3
import xlsxwriter
import tkinter.filedialog
from time import strptime
from tkinter import *
from tkinter import messagebox, simpledialog
from Multiherramienta import *

# Tiempos de espera para respuesta web
# wsh:tiempo de espera corto
# wle: tiempo de espera largo

OR_PASSWORD = ORACLE_ID.PASSWORD
JU_PASSWORD = JAVER_ID.PASSWORD
wsh = 7
wle = 7
wslp = 3


# HERRAMIENTAS VARIADAS -------------------------------------------------------------------------------------------------------------------------------------------------


def saturday():
    today = datetime.date.today()
    weekday = today.weekday()
    lastsat = today + datetime.timedelta(days = (- 2 - weekday))
    if lastsat.day < 10:
        day = "0" + str(lastsat.day)
    else:
        day = str(lastsat.day)
    if lastsat.month < 10:
        month = "0" + str(lastsat.month)
    else:
        month = str(lastsat.month)
    lastsaturday = day + "/" + month + "/" + str(lastsat.year)
    print (lastsaturday)
    return (lastsaturday)


def lastsaturday():
    today = datetime.date.today()
    weekday = today.weekday()
    lastsat = today + datetime.timedelta(days = (- 2 - weekday))
    if lastsat.day < 10:
        day = "0" + str(lastsat.day)
    else:
        day = str(lastsat.day)
    if lastsat.month < 10:
        month = "0" + str(lastsat.month)
    else:
        month = str(lastsat.month)

    lastsaturday = day + "/" + month + "/" + str(lastsat.year)
    print (lastsaturday)
    return (lastsaturday)


def EXC(SHEET, PATH):
    ''' N, A, B, C, D, E, F DEVUELVE VALORES DE A A F DADO UNA HOJA Y UNA RUTA DE EXCEL''' 
    e_d = openpyxl.load_workbook(PATH)
    AUT = e_d[SHEET]
    NUM = []
    A = []
    B = []
    C = []
    D = []
    E = []
    F = [] 
    N = 0
    for cell in AUT['A']:
        if cell.value != 0:
            NUM.append(cell.row)
            CONST = str(NUM[N])
            A.append(AUT['A' + CONST].value)
            B.append(AUT['B' + CONST].value)
            C.append(AUT['C' + CONST].value)
            D.append(AUT['D' + CONST].value)
            E.append(AUT['E' + CONST].value)
            F.append(AUT['F' + CONST].value)
            N += 1
    return N, A, B, C, D, E, F


def DRIVER(driver):
    '''DEFINE EL DRIVER A USAR
        LOS DRIVERS SON:
        1 = HIDEN FIREFOX
        2  = FIREFOX
        3 = HIDEN CHROME
        4  = CHROME'''

    path = 'C:\\ProgramData\\Autodesk\\PEX\\PRGS\\'
    # FIREFOX
    extens = 'application/vnd.hzn-3d-crossword,video/3gpp,video/3gpp2,application/vnd.mseq,application/vnd.3m.post-it-notes,application/vnd.3gpp.pic-bw-large,application/vnd.3gpp.pic-bw-small,application/vnd.3gpp.pic-bw-var,application/vnd.3gp2.tcap,application/x-7z-compressed,application/x-abiword,application/x-ace-compressed,application/vnd.americandynamics.acc,application/vnd.acucobol,application/vnd.acucorp,audio/adpcm,application/x-authorware-bin,application/x-athorware-map,application/x-authorware-seg,application/vnd.adobe.air-application-installer-package+zip,application/x-shockwave-flash,application/vnd.adobe.fxp,application/pdf,application/vnd.cups-ppd,application/x-director,applicaion/vnd.adobe.xdp+xml,application/vnd.adobe.xfdf,audio/x-aac,application/vnd.ahead.space,application/vnd.airzip.filesecure.azf,application/vnd.airzip.filesecure.azs,application/vnd.amazon.ebook,application/vnd.amiga.ami,applicatin/andrew-inset,application/vnd.android.package-archive,application/vnd.anser-web-certificate-issue-initiation,application/vnd.anser-web-funds-transfer-initiation,application/vnd.antix.game-component,application/vnd.apple.installe+xml,application/applixware,application/vnd.hhe.lesson-player,application/vnd.aristanetworks.swi,text/x-asm,application/atomcat+xml,application/atomsvc+xml,application/atom+xml,application/pkix-attr-cert,audio/x-aiff,video/x-msvieo,application/vnd.audiograph,image/vnd.dxf,model/vnd.dwf,text/plain-bas,application/x-bcpio,application/octet-stream,image/bmp,application/x-bittorrent,application/vnd.rim.cod,application/vnd.blueice.multipass,application/vnd.bm,application/x-sh,image/prs.btif,application/vnd.businessobjects,application/x-bzip,application/x-bzip2,application/x-csh,text/x-c,application/vnd.chemdraw+xml,text/css,chemical/x-cdx,chemical/x-cml,chemical/x-csml,application/vn.contact.cmsg,application/vnd.claymore,application/vnd.clonk.c4group,image/vnd.dvb.subtitle,application/cdmi-capability,application/cdmi-container,application/cdmi-domain,application/cdmi-object,application/cdmi-queue,applicationvnd.cluetrust.cartomobile-config,application/vnd.cluetrust.cartomobile-config-pkg,image/x-cmu-raster,model/vnd.collada+xml,text/csv,application/mac-compactpro,application/vnd.wap.wmlc,image/cgm,x-conference/x-cooltalk,image/x-cmx,application/vnd.xara,application/vnd.cosmocaller,application/x-cpio,application/vnd.crick.clicker,application/vnd.crick.clicker.keyboard,application/vnd.crick.clicker.palette,application/vnd.crick.clicker.template,application/vn.crick.clicker.wordbank,application/vnd.criticaltools.wbs+xml,application/vnd.rig.cryptonote,chemical/x-cif,chemical/x-cmdf,application/cu-seeme,application/prs.cww,text/vnd.curl,text/vnd.curl.dcurl,text/vnd.curl.mcurl,text/vnd.crl.scurl,application/vnd.curl.car,application/vnd.curl.pcurl,application/vnd.yellowriver-custom-menu,application/dssc+der,application/dssc+xml,application/x-debian-package,audio/vnd.dece.audio,image/vnd.dece.graphic,video/vnd.dec.hd,video/vnd.dece.mobile,video/vnd.uvvu.mp4,video/vnd.dece.pd,video/vnd.dece.sd,video/vnd.dece.video,application/x-dvi,application/vnd.fdsn.seed,application/x-dtbook+xml,application/x-dtbresource+xml,application/vnd.dvb.ait,applcation/vnd.dvb.service,audio/vnd.digital-winds,image/vnd.djvu,application/xml-dtd,application/vnd.dolby.mlp,application/x-doom,application/vnd.dpgraph,audio/vnd.dra,application/vnd.dreamfactory,audio/vnd.dts,audio/vnd.dts.hd,imag/vnd.dwg,application/vnd.dynageo,application/ecmascript,application/vnd.ecowin.chart,image/vnd.fujixerox.edmics-mmr,image/vnd.fujixerox.edmics-rlc,application/exi,application/vnd.proteus.magazine,application/epub+zip,message/rfc82,application/vnd.enliven,application/vnd.is-xpr,image/vnd.xiff,application/vnd.xfdl,application/emma+xml,application/vnd.ezpix-album,application/vnd.ezpix-package,image/vnd.fst,video/vnd.fvt,image/vnd.fastbidsheet,application/vn.denovo.fcselayout-link,video/x-f4v,video/x-flv,image/vnd.fpx,image/vnd.net-fpx,text/vnd.fmi.flexstor,video/x-fli,application/vnd.fluxtime.clip,application/vnd.fdf,text/x-fortran,application/vnd.mif,application/vnd.framemaker,imae/x-freehand,application/vnd.fsc.weblaunch,application/vnd.frogans.fnc,application/vnd.frogans.ltf,application/vnd.fujixerox.ddd,application/vnd.fujixerox.docuworks,application/vnd.fujixerox.docuworks.binder,application/vnd.fujitu.oasys,application/vnd.fujitsu.oasys2,application/vnd.fujitsu.oasys3,application/vnd.fujitsu.oasysgp,application/vnd.fujitsu.oasysprs,application/x-futuresplash,application/vnd.fuzzysheet,image/g3fax,application/vnd.gmx,model/vn.gtw,application/vnd.genomatix.tuxedo,application/vnd.geogebra.file,application/vnd.geogebra.tool,model/vnd.gdl,application/vnd.geometry-explorer,application/vnd.geonext,application/vnd.geoplan,application/vnd.geospace,applicatio/x-font-ghostscript,application/x-font-bdf,application/x-gtar,application/x-texinfo,application/x-gnumeric,application/vnd.google-earth.kml+xml,application/vnd.google-earth.kmz,application/vnd.grafeq,image/gif,text/vnd.graphviz,aplication/vnd.groove-account,application/vnd.groove-help,application/vnd.groove-identity-message,application/vnd.groove-injector,application/vnd.groove-tool-message,application/vnd.groove-tool-template,application/vnd.groove-vcar,video/h261,video/h263,video/h264,application/vnd.hp-hpid,application/vnd.hp-hps,application/x-hdf,audio/vnd.rip,application/vnd.hbci,application/vnd.hp-jlyt,application/vnd.hp-pcl,application/vnd.hp-hpgl,application/vnd.yamaha.h-script,application/vnd.yamaha.hv-dic,application/vnd.yamaha.hv-voice,application/vnd.hydrostatix.sof-data,application/hyperstudio,application/vnd.hal+xml,text/html,application/vnd.ibm.rights-management,application/vnd.ibm.securecontainer,text/calendar,application/vnd.iccprofile,image/x-icon,application/vnd.igloader,image/ief,application/vnd.immervision-ivp,application/vnd.immervision-ivu,application/reginfo+xml,text/vnd.in3d.3dml,text/vnd.in3d.spot,mode/iges,application/vnd.intergeo,application/vnd.cinderella,application/vnd.intercon.formnet,application/vnd.isac.fcs,application/ipfix,application/pkix-cert,application/pkixcmp,application/pkix-crl,application/pkix-pkipath,applicaion/vnd.insors.igm,application/vnd.ipunplugged.rcprofile,application/vnd.irepository.package+xml,text/vnd.sun.j2me.app-descriptor,application/java-archive,application/java-vm,application/x-java-jnlp-file,application/java-serializd-object,text/x-java-source,java,application/javascript,application/json,application/vnd.joost.joda-archive,video/jpm,image/jpeg,video/jpeg,application/vnd.kahootz,application/vnd.chipnuts.karaoke-mmd,application/vnd.kde.karbon,aplication/vnd.kde.kchart,application/vnd.kde.kformula,application/vnd.kde.kivio,application/vnd.kde.kontour,application/vnd.kde.kpresenter,application/vnd.kde.kspread,application/vnd.kde.kword,application/vnd.kenameaapp,applicatin/vnd.kidspiration,application/vnd.kinar,application/vnd.kodak-descriptor,application/vnd.las.las+xml,application/x-latex,application/vnd.llamagraphics.life-balance.desktop,application/vnd.llamagraphics.life-balance.exchange+xml,application/vnd.jam,application/vnd.lotus-1-2-3,application/vnd.lotus-approach,application/vnd.lotus-freelance,application/vnd.lotus-notes,application/vnd.lotus-organizer,application/vnd.lotus-screencam,application/vnd.lotus-wordro,audio/vnd.lucent.voice,audio/x-mpegurl,video/x-m4v,application/mac-binhex40,application/vnd.macports.portpkg,application/vnd.osgeo.mapguide.package,application/marc,application/marcxml+xml,application/mxf,application/vnd.wolfrm.player,application/mathematica,application/mathml+xml,application/mbox,application/vnd.medcalcdata,application/mediaservercontrol+xml,application/vnd.mediastation.cdkey,application/vnd.mfer,application/vnd.mfmp,model/mesh,appliation/mads+xml,application/mets+xml,application/mods+xml,application/metalink4+xml,application/vnd.ms-powerpoint.template.macroenabled.12,application/vnd.ms-word.document.macroenabled.12,application/vnd.ms-word.template.macroenabed.12,application/vnd.mcd,application/vnd.micrografx.flo,application/vnd.micrografx.igx,application/vnd.eszigno3+xml,application/x-msaccess,video/x-ms-asf,application/x-msdownload,application/vnd.ms-artgalry,application/vnd.ms-ca-compressed,application/vnd.ms-ims,application/x-ms-application,application/x-msclip,image/vnd.ms-modi,application/vnd.ms-fontobject,application/vnd.ms-excel,application/vnd.ms-excel.addin.macroenabled.12,application/vnd.ms-excelsheet.binary.macroenabled.12,application/vnd.ms-excel.template.macroenabled.12,application/vnd.ms-excel.sheet.macroenabled.12,application/vnd.ms-htmlhelp,application/x-mscardfile,application/vnd.ms-lrm,application/x-msmediaview,aplication/x-msmoney,application/vnd.openxmlformats-officedocument.presentationml.presentation,application/vnd.openxmlformats-officedocument.presentationml.slide,application/vnd.openxmlformats-officedocument.presentationml.slideshw,application/vnd.openxmlformats-officedocument.presentationml.template,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/vnd.openxmlformats-officedocument.spreadsheetml.template,application/vnd.openxmformats-officedocument.wordprocessingml.document,application/vnd.openxmlformats-officedocument.wordprocessingml.template,application/x-msbinder,application/vnd.ms-officetheme,application/onenote,audio/vnd.ms-playready.media.pya,vdeo/vnd.ms-playready.media.pyv,application/vnd.ms-powerpoint,application/vnd.ms-powerpoint.addin.macroenabled.12,application/vnd.ms-powerpoint.slide.macroenabled.12,application/vnd.ms-powerpoint.presentation.macroenabled.12,appliation/vnd.ms-powerpoint.slideshow.macroenabled.12,application/vnd.ms-project,application/x-mspublisher,application/x-msschedule,application/x-silverlight-app,application/vnd.ms-pki.stl,application/vnd.ms-pki.seccat,application/vn.visio,video/x-ms-wm,audio/x-ms-wma,audio/x-ms-wax,video/x-ms-wmx,application/x-ms-wmd,application/vnd.ms-wpl,application/x-ms-wmz,video/x-ms-wmv,video/x-ms-wvx,application/x-msmetafile,application/x-msterminal,application/msword,application/x-mswrite,application/vnd.ms-works,application/x-ms-xbap,application/vnd.ms-xpsdocument,audio/midi,application/vnd.ibm.minipay,application/vnd.ibm.modcap,application/vnd.jcp.javame.midlet-rms,application/vnd.tmobile-ivetv,application/x-mobipocket-ebook,application/vnd.mobius.mbk,application/vnd.mobius.dis,application/vnd.mobius.plc,application/vnd.mobius.mqy,application/vnd.mobius.msl,application/vnd.mobius.txf,application/vnd.mobius.daf,tex/vnd.fly,application/vnd.mophun.certificate,application/vnd.mophun.application,video/mj2,audio/mpeg,video/vnd.mpegurl,video/mpeg,application/mp21,audio/mp4,video/mp4,application/mp4,application/vnd.apple.mpegurl,application/vnd.msician,application/vnd.muvee.style,application/xv+xml,application/vnd.nokia.n-gage.data,application/vnd.nokia.n-gage.symbian.install,application/x-dtbncx+xml,application/x-netcdf,application/vnd.neurolanguage.nlu,application/vnd.na,application/vnd.noblenet-directory,application/vnd.noblenet-sealer,application/vnd.noblenet-web,application/vnd.nokia.radio-preset,application/vnd.nokia.radio-presets,text/n3,application/vnd.novadigm.edm,application/vnd.novadim.edx,application/vnd.novadigm.ext,application/vnd.flographit,audio/vnd.nuera.ecelp4800,audio/vnd.nuera.ecelp7470,audio/vnd.nuera.ecelp9600,application/oda,application/ogg,audio/ogg,video/ogg,application/vnd.oma.dd2+xml,applicatin/vnd.oasis.opendocument.text-web,application/oebps-package+xml,application/vnd.intu.qbo,application/vnd.openofficeorg.extension,application/vnd.yamaha.openscoreformat,audio/webm,video/webm,application/vnd.oasis.opendocument.char,application/vnd.oasis.opendocument.chart-template,application/vnd.oasis.opendocument.database,application/vnd.oasis.opendocument.formula,application/vnd.oasis.opendocument.formula-template,application/vnd.oasis.opendocument.grapics,application/vnd.oasis.opendocument.graphics-template,application/vnd.oasis.opendocument.image,application/vnd.oasis.opendocument.image-template,application/vnd.oasis.opendocument.presentation,application/vnd.oasis.opendocumen.presentation-template,application/vnd.oasis.opendocument.spreadsheet,application/vnd.oasis.opendocument.spreadsheet-template,application/vnd.oasis.opendocument.text,application/vnd.oasis.opendocument.text-master,application/vnd.asis.opendocument.text-template,image/ktx,application/vnd.sun.xml.calc,application/vnd.sun.xml.calc.template,application/vnd.sun.xml.draw,application/vnd.sun.xml.draw.template,application/vnd.sun.xml.impress,application/vnd.sun.xl.impress.template,application/vnd.sun.xml.math,application/vnd.sun.xml.writer,application/vnd.sun.xml.writer.global,application/vnd.sun.xml.writer.template,application/x-font-otf,application/vnd.yamaha.openscoreformat.osfpvg+xml,application/vnd.osgi.dp,application/vnd.palm,text/x-pascal,application/vnd.pawaafile,application/vnd.hp-pclxl,application/vnd.picsel,image/x-pcx,image/vnd.adobe.photoshop,application/pics-rules,image/x-pict,application/x-chat,aplication/pkcs10,application/x-pkcs12,application/pkcs7-mime,application/pkcs7-signature,application/x-pkcs7-certreqresp,application/x-pkcs7-certificates,application/pkcs8,application/vnd.pocketlearn,image/x-portable-anymap,image/-portable-bitmap,application/x-font-pcf,application/font-tdpfr,application/x-chess-pgn,image/x-portable-graymap,image/png,image/x-portable-pixmap,application/pskc+xml,application/vnd.ctc-posml,application/postscript,application/xfont-type1,application/vnd.powerbuilder6,application/pgp-encrypted,application/pgp-signature,application/vnd.previewsystems.box,application/vnd.pvi.ptid1,application/pls+xml,application/vnd.pg.format,application/vnd.pg.osasli,tex/prs.lines.tag,application/x-font-linux-psf,application/vnd.publishare-delta-tree,application/vnd.pmi.widget,application/vnd.quark.quarkxpress,application/vnd.epson.esf,application/vnd.epson.msf,application/vnd.epson.ssf,applicaton/vnd.epson.quickanime,application/vnd.intu.qfx,video/quicktime,application/x-rar-compressed,audio/x-pn-realaudio,audio/x-pn-realaudio-plugin,application/rsd+xml,application/vnd.rn-realmedia,application/vnd.realvnc.bed,applicatin/vnd.recordare.musicxml,application/vnd.recordare.musicxml+xml,application/relax-ng-compact-syntax,application/vnd.data-vision.rdz,application/rdf+xml,application/vnd.cloanto.rp9,application/vnd.jisp,application/rtf,text/richtex,application/vnd.route66.link66+xml,application/rss+xml,application/shf+xml,application/vnd.sailingtracker.track,image/svg+xml,application/vnd.sus-calendar,application/sru+xml,application/set-payment-initiation,application/set-reistration-initiation,application/vnd.sema,application/vnd.semd,application/vnd.semf,application/vnd.seemail,application/x-font-snf,application/scvp-vp-request,application/scvp-vp-response,application/scvp-cv-request,application/svp-cv-response,application/sdp,text/x-setext,video/x-sgi-movie,application/vnd.shana.informed.formdata,application/vnd.shana.informed.formtemplate,application/vnd.shana.informed.interchange,application/vnd.shana.informed.package,application/thraud+xml,application/x-shar,image/x-rgb,application/vnd.epson.salt,application/vnd.accpac.simply.aso,application/vnd.accpac.simply.imp,application/vnd.simtech-mindmapper,application/vnd.commonspace,application/vnd.ymaha.smaf-audio,application/vnd.smaf,application/vnd.yamaha.smaf-phrase,application/vnd.smart.teacher,application/vnd.svd,application/sparql-query,application/sparql-results+xml,application/srgs,application/srgs+xml,application/sml+xml,application/vnd.koan,text/sgml,application/vnd.stardivision.calc,application/vnd.stardivision.draw,application/vnd.stardivision.impress,application/vnd.stardivision.math,application/vnd.stardivision.writer,application/vnd.tardivision.writer-global,application/vnd.stepmania.stepchart,application/x-stuffit,application/x-stuffitx,application/vnd.solent.sdkm+xml,application/vnd.olpc-sugar,audio/basic,application/vnd.wqd,application/vnd.symbian.install,application/smil+xml,application/vnd.syncml+xml,application/vnd.syncml.dm+wbxml,application/vnd.syncml.dm+xml,application/x-sv4cpio,application/x-sv4crc,application/sbml+xml,text/tab-separated-values,image/tiff,application/vnd.to.intent-module-archive,application/x-tar,application/x-tcl,application/x-tex,application/x-tex-tfm,application/tei+xml,text/plain,application/vnd.spotfire.dxp,application/vnd.spotfire.sfs,application/timestamped-data,applicationvnd.trid.tpt,application/vnd.triscape.mxs,text/troff,application/vnd.trueapp,application/x-font-ttf,text/turtle,application/vnd.umajin,application/vnd.uoml+xml,application/vnd.unity,application/vnd.ufdl,text/uri-list,application/nd.uiq.theme,application/x-ustar,text/x-uuencode,text/x-vcalendar,text/x-vcard,application/x-cdlink,application/vnd.vsf,model/vrml,application/vnd.vcx,model/vnd.mts,model/vnd.vtu,application/vnd.visionary,video/vnd.vivo,applicatin/ccxml+xml,,application/voicexml+xml,application/x-wais-source,application/vnd.wap.wbxml,image/vnd.wap.wbmp,audio/x-wav,application/davmount+xml,application/x-font-woff,application/wspolicy+xml,image/webp,application/vnd.webturb,application/widget,application/winhlp,text/vnd.wap.wml,text/vnd.wap.wmlscript,application/vnd.wap.wmlscriptc,application/vnd.wordperfect,application/vnd.wt.stf,application/wsdl+xml,image/x-xbitmap,image/x-xpixmap,image/x-xwindowump,application/x-x509-ca-cert,application/x-xfig,application/xhtml+xml,application/xml,application/xcap-diff+xml,application/xenc+xml,application/patch-ops-error+xml,application/resource-lists+xml,application/rls-services+xml,aplication/resource-lists-diff+xml,application/xslt+xml,application/xop+xml,application/x-xpinstall,application/xspf+xml,application/vnd.mozilla.xul+xml,chemical/x-xyz,text/yaml,application/yang,application/yin+xml,application/vnd.ul,application/zip,application/vnd.handheld-entertainment+xml,application/vnd.zzazz.deck+xml'
    fp1 = Options()
    fp1.add_argument('-headless')
    fp1.set_preference("browser.download.folderList", 2)
    fp1.set_preference("browser.download.dir", path)
    fp1.set_preference("browser.download.useDownloadDir", True)
    fp1.set_preference("browser.download.manager.showWhenStarting", False)
    fp1.set_preference("browser.helperApps.alwaysAsk.force", False)
    fp1.set_preference("browser.helperApps.neverAsk.saveToDisk", extens)
    fp1.set_preference("browser.download.manager.showAlertOnComplete", False)
    fp1.set_preference("browser.download.manager.useWindow", False)
    fp1.set_preference("pdfjs.disabled", True)
    fp1.set_preference("plugin.scan.plid.all", False)
    fp1.set_preference("dom.popup_maximum", 100)
    fp = Options()
    fp.set_preference("browser.download.folderList", 2)
    fp.set_preference("browser.download.dir", path)
    fp.set_preference("browser.download.useDownloadDir", True)
    fp.set_preference("browser.download.manager.showWhenStarting", False)
    fp.set_preference("browser.helperApps.alwaysAsk.force", False)
    fp.set_preference("browser.helperApps.neverAsk.saveToDisk", extens)
    fp.set_preference("browser.download.manager.showAlertOnComplete", False)
    fp.set_preference("browser.download.manager.useWindow", False)
    fp.set_preference("pdfjs.disabled", True)
    fp.set_preference("plugin.scan.plid.all", False)
    fp.set_preference("dom.popup_maximum", 100)
    GECKODV = Service("C:\\ProgramData\\Autodesk\\PEX\\PRGS\\geckodriver.exe")
    CHRDV = Service("C:\\ProgramData\\Autodesk\\PEX\\PRGS\\chromedriver.exe")
    # CHROME
    CHROME_OPTIONS = webdriver.ChromeOptions()
    CHROME_OPTIONS.add_argument("--headless")
    if driver == 1:
        DRIVE = webdriver.Firefox(options=fp1, service=GECKODV)
    elif driver == 2:
        DRIVE = webdriver.Firefox(options=fp, service=GECKODV)
    elif driver == 3:
        DRIVE = webdriver.Chrome(service=CHRDV, options=CHROME_OPTIONS)
    elif driver == 4:
        DRIVE = webdriver.Chrome(service=CHRDV)
    return DRIVE


def CONJUNTO (CONJUNTO):
    ''' DADO UN CONJUNTO DEVUELVE EL: FRACCIONAMIENTO, FRENTE, ETAPA'''
    CON = ["E01", "E02", "E03", "E04", "E05", "E06", "E07", "E08", "E09", "E10", "E11", "E12", "E13", "E14", "E15", "E16", "E17", "E18", "E19", "E20", "E21", "E22", "E23", "E24", "E25", "E26", "E27", "E28"]
    UJN = ["VSU", "   ", "RLU", "BSU", "S2U", "BDU", "P2U", "   ", "   ", "USV", "ELU", "UED", "UB4", "   ", "URC", "UNL", "JUU", "USI", "   ", "UR7", "UMA", "UFB", "UPR", "UMO", "CJM", "   ", "   ", "   "]
    CJQ = ["CST", "   ", "CRL", "CB2", "CS2", "CB3", "CHM", "   ", "   ", "CS3", "CEL", "CVP", "CB4", "   ", "CRÑ", "CMN", "CBJ", "CSI", "   ", "CR7", "CMA", "CFB", "CPR", "CMS", "CJM", "   ", "PME", "VDV"]
    #SCJ = 
    PRIMER = CONJUNTO[0:3]
    ETAPA = CONJUNTO[7:10]
    FRENTE = CONJUNTO[4:6]
    NUMERO = int(PRIMER[1:3]) - 1
    if ETAPA == "I01" or ETAPA == "U02":
        if int(FRENTE) > 79:
            FRAC = CJQ[NUMERO]
        else:
            FRAC = UJN[NUMERO]
    else: 
        FRAC = CJQ[NUMERO]
    return (FRAC, FRENTE, ETAPA)


#HERRAMIENTAS ORACLE-----------------------------------------------------------------------------------------------------------------------------------------------------


def acc_fir_oracle(driver):
    # ACCEDE A ORACLE CON LA CUENTA FPRADO
    driver.get("http://siapp3.javer.com.mx:8010/OA_HTML/AppsLogin")
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.NAME, "usernameField"))).clear()
    WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.NAME, "usernameField"))).send_keys("FPRADO")
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.NAME, "passwordField"))).send_keys(OR_PASSWORD + Keys.ENTER)


def encontrar_lista(driver, INI, ITEM, ID_I, ID_F):
    LISTA = []
    time.sleep(0.5)
    x = 0
    while True:
        if x == ITEM:
            break
        try:
            A = str(x + INI)
            SELECTOR = str(ID_I) + A + str(ID_F)
            LISTA.append(WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.CSS_SELECTOR, SELECTOR))).text)
            x += 1
        except:
            break
    return LISTA


def encontrar_a_sr(driver, donde):
    # ENCUENTRA UN REPORTE EN EL SISTEMA DE REPORTE (DRIVER, DONDE = NOMBRE DEL REPORTE)
    REPORTE = []
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="buttonRefreshReport"]')))
    for N in range(17):
        CONS = str(N)
        XXPATH = '//*[@id="TableReportsRN:itemName:' + CONS + '"]'
        REPORTE.append(WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, XXPATH))).text)
    FILA = str(REPORTE.index(donde))
    return FILA


def objetivo(driver, que, donde):
    ''' DESCARGA UN ARCHIVO EN EL MENÚ DE SISTEMA DE REPORTES DONDE
        que = FORMATO (PDF, XLS, ETC.)
        donde = Nombre del reporte'''
    D = encontrar_a_sr(driver, donde)
    BOTT = '//*[@id="TableReportsRN:' + que + 'D:' + D + '"]'
    STAT = WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, BOTT))).get_attribute('title')
    while STAT == 'Salida actualmente no disponible':
        try:
            STAT = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, BOTT))).get_attribute('title')
            WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="buttonRefreshReport"]'))).click()
        except StaleElementReferenceException:
            WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="buttonRefreshReport"]'))).click()
        except TimeoutException:
            try:
                BOTT = '//*[@id="TableReportsRN:' + que + 'Wimg:' + D + '"]'
                STAT = WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, BOTT))).get_attribute('id')
            except TimeoutException:
                BOTT = '//*[@id="TableReportsRN:' + que + '0img:' + D + '"]'
                WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, BOTT))).click()
                BOTT = '//*[@id="TableReportsRN:' + que + 'Wimg:' + D + '"]'
                STAT = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, BOTT))).get_attribute('id')
            except StaleElementReferenceException:
                WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="buttonRefreshReport"]'))).click()
                WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, BOTT))).click()
    while STAT == 'TableReportsRN:' + que + 'Wimg:' + D:
        try:
            WebDriverWait(driver, 3).until(EC.presence_of_element_located((By.XPATH, BOTT))).click()
        except TimeoutException:
            BOTT = '//*[@id="TableReportsRN:' + que + '1img:' + D + '"]'
            STAT = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, BOTT))).get_attribute('id')
            WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, BOTT))).click()
        except StaleElementReferenceException:
            WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="buttonRefreshReport"]'))).click()
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, BOTT))).click()
        except:
            continue


def frame_click_go(driver, value):
    ''' Ingresa al frame de Oracle y escrive un valor (value) en el frame 
    (DRIVER, VALOR A ESCRIBIR)'''
    TS = time.sleep (0.5)
    count = 0
    while True:
        if count == 3:
            print("Falló", end="|")
            break
        try:
            time.sleep(0.5)
            print("H", count, sep="", end="|")
            driver.switch_to.window(driver.window_handles[1])
            driver.switch_to.frame(0)
            break
        except:
            count += 1
            time.sleep(1)
            driver.switch_to.window(driver.window_handles[1])
            continue
    while True:
        if count == 3:
            print("falló", end="|")
            break
        try:
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Término de Búsqueda"]'))).clear()
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Término de Búsqueda"]'))).send_keys(value + Keys.TAB)
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "(//form[@id='_LOVResFrm']//table)[3]/tbody[1]/tr[1]/td[1]/div[1]/div[1]/button[1]"))).click()
            TS
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//table[@class='x1o']/tbody[1]/tr[2]/td[2]/a[1]/img[1]"))).click()
            print("1", end="|")
            break
        except StaleElementReferenceException:
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Término de Búsqueda"]'))).clear()
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Término de Búsqueda"]'))).send_keys(value + Keys.TAB)
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "(//form[@id='_LOVResFrm']//table)[3]/tbody[1]/tr[1]/td[1]/div[1]/div[1]/button[1]"))).click()
            TS
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//table[@class='x1o']/tbody[1]/tr[2]/td[2]/a[1]/img[1]"))).click()
            print("2", end="|")
            break
        except TimeoutException:
            try: 
                WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Término de Búsqueda"]'))).clear()
                WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Término de Búsqueda"]'))).send_keys(value + Keys.TAB + Keys.ENTER)
                WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "(//form[@id='_LOVResFrm']//table)[3]/tbody[1]/tr[1]/td[1]/div[1]/div[1]/button[1]"))).click()
                TS
                WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//table[@class='x1o']/tbody[1]/tr[2]/td[2]/a[1]/img[1]"))).click()
                print("3", end="|" )
                break
            except:
                count += 1
                driver.close()
                driver.switch_to.window(driver.window_handles[0])
                raise Exception("No hay valores")
                print("4 : Fail", end="|")
                break
        except:
            count += 1
            continue
    time.sleep(1)
    driver.switch_to.window(driver.window_handles[0])
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Logo Oracle"]')))
    time.sleep(1)
    print("", end = "")


# DEFINICIONES IR EN ORACLE----------------------------------------------------------------------------------------------------------------------------------------------


def ir_PLANTILLA(driver):
    try:
        WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="SFAV"]'))).click()
        WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/span[2]/table[1]/tbody/tr[2]/td/table/tbody/tr[2]/td[2]/table/tbody/tr/td[1]/div[3]/div/ul/div/div/div/li[15]/a'))).click()
    except TimeoutException:
        try:
            WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="SFAV"]'))).click()
            WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/span[2]/table[1]/tbody/tr[2]/td/table/tbody/tr[2]/td[2]/table/tbody/tr/td[1]/div[3]/div/ul/div/div/div/li[15]/a'))).click()
        except TimeoutException:
            WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="SFAV"]'))).click()
            WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/span[2]/table[1]/tbody/tr[2]/td/table/tbody/tr[2]/td[2]/table/tbody/tr/td[1]/div[3]/div/ul/div/div/div/li[15]/a'))).click()


def ir_en_fav(driver, nombre):
    # ACCEDE AL SISTEMA DE REPORTES UNA VEZ YA INGRESADO A ORACLE
    try:
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="SFAV"]'))).click()
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.LINK_TEXT, nombre))).click()
    except TimeoutException:
        try:
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="SFAV"]'))).click()
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.LINK_TEXT, nombre))).click()
        except TimeoutException:
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="SFAV"]'))).click()
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.LINK_TEXT, nombre))).click()


def ir_repedocue(driver):
    # ACCEDE AL MENÚ DE REPORTE DE ESTADO DE CUENTA
    try:
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="SFAV"]'))).click()
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.LINK_TEXT, 'RECC'))).click()
    except TimeoutException:
        try:
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="SFAV"]'))).click()
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.LINK_TEXT, 'RECC'))).click()
        except TimeoutException:
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="SFAV"]'))).click()
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.LINK_TEXT, 'RECC'))).click()


def ir_finob(driver):
    JMCQ = "//table[@id='respList']/tbody[1]/tr[1]/td[1]/ul[1]/li[1]/a[1]/img[1]"
    JMRG = "//table[@id='respList']/tbody[1]/tr[1]/td[1]/ul[1]/li[1]/ul[1]/li[10]/a[1]/img[1]"
    RFDO = "//table[@id='respList']/tbody[1]/tr[1]/td[1]/ul[1]/li[1]/ul[1]/li[10]/ul[1]/li[22]/a[1]"
    while True:
        DONDE = 0
        try:
            DONDE = WebDriverWait(driver, 0.5).until(EC.element_to_be_clickable((By.ID, "Fndcpprogramnamedisplay"))).text
        except:
            pass
        try:
            DONDE = WebDriverWait(driver, 0.5).until(EC.element_to_be_clickable((By.XPATH, "//h1[text()='Oracle Applications Home Page']"))).text
        except:
            pass
        try:
            DONDE = WebDriverWait(driver, 0.5).until(EC.element_to_be_clickable((By.XPATH, "//h1[text()='Contratos']"))).text
        except:
            pass
        try:
            DONDE = WebDriverWait(driver, 0.5).until(EC.element_to_be_clickable((By.XPATH, "//h1[text()='Requests']"))).text
        except:
            pass
        if DONDE == "XXMCAN - Finiquito de Obra":
            break
        elif DONDE == "Contratos" or DONDE =="Requests":
            WebDriverWait(driver, 0.5).until(EC.element_to_be_clickable((By.XPATH, "(//table[@cellspacing='0']//a)[4]"))).click()
        elif DONDE == "Oracle Applications Home Page":
            CHEK0 = WebDriverWait(driver, 0.5).until(EC.element_to_be_clickable((By.XPATH, JMCQ))).get_attribute("alt")
            if CHEK0 == "Expand":
                WebDriverWait(driver, 0.5).until(EC.element_to_be_clickable((By.XPATH, JMCQ))).click()
            CHEK1 = WebDriverWait(driver, 0.5).until(EC.element_to_be_clickable((By.XPATH, JMRG))).get_attribute("alt")
            if CHEK1 == "Expand":
                WebDriverWait(driver, 0.5).until(EC.element_to_be_clickable((By.XPATH, JMRG))).click()
            WebDriverWait(driver, 0.5).until(EC.element_to_be_clickable((By.XPATH, RFDO))).click()


def ir_en_rg(driver, donde):
    D = encontrar_a_sr(driver, donde)
    try:
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="TableReportsRN:ICOEXE1img:' + D + '"]'))).click()
    except TimeoutException:
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="TableReportsRN:ICOEXE1img:' + D + '"]'))).click()
    time.sleep(0.5)


# SUBPROGRAMAS ORACLE ---------------------------------------------------------------------------------------------------------------------------------------------------


def datos_recc2(driver,proy,fren):
    # En el Reporte de finiquito de obra ingresa datos de proyecto y frente
    while True:
        try:
            ir_finob(driver)
            WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.ID, "Fndcpprogramnamedisplay")))
            break
        except:
            continue
    while True:
        try:
            SIGUIENTE = "(//table[@class='x6w']//table)[3]/tbody[1]/tr[1]/td[3]/button[1]"
            WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, SIGUIENTE))).click()
            break
        except:
            continue
    while True:
        try:
            WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="N610"]'))).send_keys(Keys.CLEAR + Keys.CLEAR + Keys.CLEAR + Keys.BACKSPACE + Keys.BACKSPACE + Keys.BACKSPACE)        
            WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="N610"]'))).send_keys(proy + Keys.TAB + Keys.TAB + Keys.TAB + Keys.TAB)
            WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, "//table[@id='Fndcpparameterregion']//table/tbody[1]/tr[2]/td[2]/span[1]")))
            break
        except:
            continue
    while True:
        try:
            WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="N611"]'))).send_keys(Keys.CLEAR + Keys.CLEAR + Keys.BACKSPACE + Keys.BACKSPACE)
            WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="N611"]'))).send_keys(fren + Keys.TAB + Keys.TAB + Keys.TAB + Keys.TAB)
            WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, "//table[@id='Fndcpparameterregion']//table/tbody[1]/tr[4]/td[2]/span[1]")))
            break
        except:
            continue
    while True:
        try:
            WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="N613"]'))).send_keys(Keys.CLEAR + Keys.CLEAR + Keys.BACKSPACE + Keys.BACKSPACE)
            WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="N613"]'))).send_keys(Keys.BACKSPACE + Keys.BACKSPACE)
            WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="N613"]'))).send_keys("16" + Keys.TAB + Keys.TAB + Keys.TAB + Keys.TAB)
            WebDriverWait(driver, 2).until(EC.presence_of_element_located((By.XPATH, "//span[text()='IVA 16%']")))
            break
        except:
            continue      
    while True:
        try:
            WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="FndReqSubmit"]'))).click()
            WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.XPATH, "(//tr[@id='Fndcpprogramname__xc_']//span)[1]")))
            break
        except:
            continue
    while True:
        try:
            WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="FndReqSubmit"]'))).click()
            break
        except:
            continue
    while True:
        try:
            inform = '/html/body/form/span[2]/div/div[3]/div/div[2]/table[2]/tbody/tr/td'
            A = WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, inform))).text
            WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, "//button[@type='button']"))).click()
            print(A[-8:])
            break
        except:
            continue
    return A


def datos_recc(driver, proy, fren):
    # Dentro del Reporte de estado de cuenta de contratos ingresa los datos de proyecto y frente
    while True:
        try:
            ir_repedocue(driver)
            time.sleep(0.5)
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "Fndcpprogramnamedisplay")))
            SIGUIENTE = '/html/body/form/span[2]/div/div[3]/div/div[2]/table/tbody/tr[1]/td[2]/table/tbody/tr/td[2]/table/tbody/tr/td/table/tbody/tr/td[4]/table/tbody/tr/td[3]/button'
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, SIGUIENTE))).click()
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="N610"]'))).send_keys(proy + Keys.TAB + Keys.TAB + Keys.TAB + Keys.TAB)
            time.sleep(1)
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="N611"]'))).click()
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="N611"]'))).send_keys(fren)
            WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="N611"]'))).send_keys(Keys.TAB)
            time.sleep(1)
            WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="FndReqSubmit"]'))).click()
            time.sleep(1)
            WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="FndReqSubmit"]'))).click()
            time.sleep(1)
            inform = '/html/body/form/span[2]/div/div[3]/div/div[2]/table[2]/tbody/tr/td'
            A = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, inform))).text
            time.sleep(0.5)
            ACEPTAR = '/html/body/form/span[2]/div/div[3]/div[2]/table/tbody/tr[1]/td[2]/table/tbody/tr/td[2]/button'
            WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, ACEPTAR))).click()
            break
        except:
            continue
    return A


def ecuentra_estimaciones(driver, proy, frent, conj, prov, prefecha):
    # Dentro del Reporte de estimaciones para contratista ingresa los datos de proyecto, frente, conjunto, provedor, fecha
    FECHAF = []
    count = 1
    while True:
        if count == 3:
            print("Falló")
            break
        try:
            PROYECTO =  "/html/body/form/span[2]/div/div[3]/div[1]/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr[1]/td[3]/span/a/img"
            FRENTE =    "/html/body/form/span[2]/div/div[3]/div[1]/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr[2]/td[3]/span/a/img"
            CONJUNTO =  "/html/body/form/span[2]/div/div[3]/div[1]/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr[3]/td[3]/span/a/img"
            PROVEEDOR = "/html/body/form/span[2]/div/div[3]/div[1]/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr[4]/td[3]/span/a/img"
            ORDENC =    "/html/body/form/span[2]/div/div[3]/div[1]/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr[5]/td[3]/span/a/img"
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, PROYECTO))).click()
            frame_click_go(driver, proy)
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Nombre del frente"]'))).click()
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, FRENTE))).click()
            frame_click_go(driver, frent)
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Nombre del conjunto"]'))).click()
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, CONJUNTO))).click()
            frame_click_go(driver, conj)
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Nombre del proveedor"]'))).click()
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, PROVEEDOR))).click()
            frame_click_go(driver, prov)
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Orden de compra de la estimacion"]'))).send_keys('%')
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, ORDENC))).click()
            break
        except:
            time.sleep(1.5)
            count += 1
            continue
    x = 0
    while True:
        if x == 10:
            break
        while True:
            if count == 3:
                print("Falló handle", end="|")
                break
            try:
                time.sleep(1)
                driver.switch_to.window(driver.window_handles[1])
                driver.switch_to.frame(0)
                break
            except NoSuchFrameException:
                continue
            except:
                count += 1
                time.sleep(1)
                driver.switch_to.window(driver.window_handles[1])
                driver.switch_to.frame(0)
                continue
        try:
            V = str(x)
            FDATO = "N1:displayColumn4:" + V
            FECHA = strptime(prefecha, "%d/%m/%Y")
            PREFECHA = strptime(WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.ID, FDATO))).text, "%Y-%m-%d %H:%M:%S.%f")
            x += 1
            if PREFECHA >= FECHA:
                PREOC = "N1:displayColumn3:" + V
                FECHAF.append(WebDriverWait(driver, 1).until(EC.presence_of_element_located((By.ID, PREOC))).text)
            else:
                break
        except StaleElementReferenceException:
            print ("FAILED")
            break
        except:
                x += 1
                time.sleep(0.2)
                driver.switch_to.window(driver.window_handles[1])
                driver.switch_to.frame(0)
                continue
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, '/html/body/span/div[1]/div[2]/table/tbody/tr/td[3]/table/tbody/tr/td[2]/button'))).click()
    time.sleep(1)
    try: 
        driver.switch_to.window(driver.window_handles[0])
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Logo Oracle"]')))
    except:
        driver.switch_to.window(driver.window_handles[0])
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Logo Oracle"]')))   
    time.sleep(1)
    ir_en_fav(driver, 'Sistema de reportes')
    return FECHAF


def datos_repepc(driver, proy, frent, conj, prov, prefecha, oc):
    # Dentro del Reporte de estimaciones para contratista ingresa los datos de proyecto, frente, conjunto, provedor, fecha
    count = 1
    while True:
        if count == 3:
            print("Falló")
            break
        try:
            PROYECTO =  "/html/body/form/span[2]/div/div[3]/div[1]/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr[1]/td[3]/span/a/img"
            FRENTE =    "/html/body/form/span[2]/div/div[3]/div[1]/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr[2]/td[3]/span/a/img"
            CONJUNTO =  "/html/body/form/span[2]/div/div[3]/div[1]/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr[3]/td[3]/span/a/img"
            PROVEEDOR = "/html/body/form/span[2]/div/div[3]/div[1]/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr[4]/td[3]/span/a/img"
            ORDENC =    "/html/body/form/span[2]/div/div[3]/div[1]/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr[5]/td[3]/span/a/img"
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, PROYECTO))).click()
            print("E", end=":")
            frame_click_go(driver, proy)
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Nombre del frente"]'))).click()
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, FRENTE))).click()
            print("F", end=":")
            frame_click_go(driver, frent)
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Nombre del conjunto"]'))).click()
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, CONJUNTO))).click()
            print("C", end=":")
            frame_click_go(driver, conj)
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Nombre del proveedor"]'))).click()
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, PROVEEDOR))).click()
            print("P", end=":")
            frame_click_go(driver, prov)
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Orden de compra de la estimacion"]'))).click()
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, ORDENC))).click()
            print("O", end=":")
            frame_click_go(driver, oc)
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "ButtonExecute"))).click()
            break
        except:
            time.sleep(1.5)
            count += 1
            continue


def datos_repcdvs(driver, proy, frent, conj, prov, finicio, ffinal):
    # Dentro del Reporte de estimación con detalle de vivienda sembrado ingresa poryecto, frente, conjunto, provedor y fecha de estimación
    CDREPCDVS = 0
    while True:
        if CDREPCDVS == 2:
            print ("falló")
            break
        try:
            PROYECTO =  "/html/body/form/span[2]/div/div[3]/div[1]/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr[1]/td[3]/span/a/img"
            FRENTE =    "/html/body/form/span[2]/div/div[3]/div[1]/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr[2]/td[3]/span/a/img"
            CONJUNTO =  "/html/body/form/span[2]/div/div[3]/div[1]/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr[3]/td[3]/span/a/img"
            PROVEEDOR = "/html/body/form/span[2]/div/div[3]/div[1]/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr[4]/td[3]/span/a/img"
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, PROYECTO))).click()
            print("E", end=":")
            frame_click_go(driver, proy)
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Nombre del frente"]'))).click()
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, FRENTE))).click()
            print("F", end=":")
            frame_click_go(driver, frent)
            # WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Nombre del conjunto"]'))).click()
            # WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, CONJUNTO))).click()
            # print("C", end=":")
            # frame_click_go(driver, conj)
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Nombre del proveedor"]'))).click()
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, PROVEEDOR))).click()
            print("P", end=":")
            frame_click_go(driver, prov)
            WebDriverWait(driver, wle).until(EC.element_to_be_clickable((By.NAME, "START_DATE"))).send_keys(finicio + Keys.TAB + Keys.TAB + Keys.TAB + Keys.TAB)
            time.sleep(.5)
            WebDriverWait(driver, wle).until(EC.element_to_be_clickable((By.NAME, "START_DATE"))).send_keys(Keys.TAB + Keys.TAB + Keys.TAB + Keys.TAB)
            print("H", end=":")
            if ffinal != "":
                WebDriverWait(driver, wle).until(EC.element_to_be_clickable((By.NAME, "END_DATE"))).send_keys(ffinal + Keys.TAB + Keys.TAB + Keys.TAB + Keys.TAB)
                time.sleep(.5)
                WebDriverWait(driver, wle).until(EC.element_to_be_clickable((By.NAME, "END_DATE"))).send_keys(Keys.TAB + Keys.TAB + Keys.TAB + Keys.TAB)
                print("H", end=":")
            time.sleep(wslp * 0.25)
            WebDriverWait(driver, wle).until(EC.element_to_be_clickable((By.NAME, "CONTRACT_OP_HDR_ID"))).send_keys(Keys.TAB + Keys.TAB + Keys.TAB + Keys.TAB + Keys.ENTER)
            WebDriverWait(driver, wle).until(EC.element_to_be_clickable((By.ID, "buttonRefreshReport")))
            break
        except:
            CDREPCDVS += 1
            print("")
            print ("    ", CDREPCDVS, end=":")
            continue


def datos_rmcc(driver, proy, frent, conj, prov):
    # Dentro del Reporte de Matriz de Contrato para Contratista ingresa los datos de Proyecto, Frente, Conjunto y Proveedor
    WebDriverWait(driver, wle).until(EC.element_to_be_clickable((By.NAME, "PROJECT_ID"))).send_keys(proy + Keys.TAB + Keys.TAB + Keys.TAB + Keys.TAB)
    time.sleep(wslp)
    WebDriverWait(driver, wle).until(EC.element_to_be_clickable((By.NAME, "PROJECT_ID"))).send_keys(Keys.TAB + Keys.TAB + Keys.TAB + Keys.TAB)
    WebDriverWait(driver, wle).until(EC.element_to_be_clickable((By.NAME, "FRONT_ID"))).send_keys(frent + Keys.TAB + Keys.TAB + Keys.TAB + Keys.TAB)
    time.sleep(wslp * 0.5)
    WebDriverWait(driver, wle).until(EC.element_to_be_clickable((By.NAME, "FRONT_ID"))).send_keys(Keys.TAB + Keys.TAB + Keys.TAB + Keys.TAB)
    WebDriverWait(driver, wle).until(EC.element_to_be_clickable((By.NAME, "FRONT_BUILD_SET_ID"))).send_keys(conj + Keys.TAB + Keys.TAB + Keys.TAB + Keys.TAB)
    time.sleep(wslp * 0.5)
    WebDriverWait(driver, wle).until(EC.element_to_be_clickable((By.NAME, "FRONT_BUILD_SET_ID"))).send_keys(Keys.TAB + Keys.TAB + Keys.TAB + Keys.TAB)
    if prov != "%":
        WebDriverWait(driver, wle).until(EC.element_to_be_clickable((By.NAME, "VENDOR_ID"))).send_keys(prov + Keys.TAB + Keys.TAB + Keys.TAB + Keys.TAB)
        time.sleep(wslp * 0.75)
        WebDriverWait(driver, wle).until(EC.element_to_be_clickable((By.NAME, "VENDOR_ID"))).send_keys(Keys.TAB + Keys.TAB + Keys.TAB + Keys.TAB)
    WebDriverWait(driver, wle).until(EC.element_to_be_clickable((By.NAME, "PACKAGE_CODE"))).send_keys(Keys.TAB + Keys.TAB + Keys.TAB + Keys.TAB + Keys.TAB + Keys.ENTER)
    WebDriverWait(driver, wle).until(EC.element_to_be_clickable((By.ID, "buttonRefreshReport")))


def descargarecc(driver, C):
    # De la lista de reportes requisitados busca y descarga los reportes de estado de cuenta ingresados en la lista
    ir_repedocue(driver)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "FndReqCancel"))).click()
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "Refresh"))).click()
    while True:
        print(len(C), type(C), C)
        if len(C) == 0:
            print("se acabó")
            break
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "Fndcpsearchquery")))
        try:
            IDD = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.ID, "N44:FndRequestId:0"))).text
            OUT = 'N44'
        except TimeoutException:
            IDD = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.ID, "N43:FndRequestId:0"))).text
            OUT = 'N43'
        for x in range(15):
            while True:
                y = str(x)
                ID = OUT + ':FndRequestId:' + y
                while  True:
                    try:
                        IDD = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, ID))).text
                        print(x + 1, IDD, type(IDD))
                        break
                    except:
                        continue
                if IDD in C or int(IDD) in C:
                    print(ID, "-", IDD)
                    # try:
                    OUTPUT = OUT + ":item1:" + y
                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, OUTPUT))).click()
                    print("ÉXITO", IDD, len(C))
                    try:
                        C.remove(IDD)
                    except:
                        C.remove(int(IDD))
                    break
                    # except:
                    #     continue
                else:
                    break
        while True:
            try:
                WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.LINK_TEXT, 'Siguientes 15'))).click()
                print("-  Siguientes 15  -")
                time.sleep(3)
                break
            except:
                WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, "//table[@class='x1p']//table/tbody[1]/tr[1]/td[5]/select[1]"))).click()
                YEAR = Select(driver.find_element_by_xpath("//table[@class='x1p']//table/tbody[1]/tr[1]/td[5]/select[1]"))
                YEAR.select_by_visible_text("1-15")
                continue
    time.sleep(5)


# SUBPROGRAMAS JURÍDICO -------------------------------------------------------------------------------------------------------------------------------------------------


def modificatorio(driver, pagina, monto, fecha, fcha2, juridico):
    # EL PROGRAMA INGRESAA LA PÁGINA
    print(pagina)
    driver.get(pagina)
    # OBTIENE EL TÍTULO DEL CONTRATO PARA VERIFICAR QUE SI SEA EL CONTRATO A MODIFICAR
    try:
        contrato = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'titleContrato'))).text
        HREF = "a[href*='/juridico/Contratos_Juridico/" + contrato + "/" + contrato + "_CF.aspx']"
        FINIQ = driver.find_element_by_css_selector(HREF).text
        print ("YA FINIQUITADO", FINIQ)
    except NoSuchElementException:
        # INTENTA OBTENER UN MODIFICATORIO YA EXISTENTE, PARA EN DICHO CASO SALTAR
        try:
            LFCH = "//table[@id='contentDS']/tbody[1]/tr[3]/td[1]/a[1]"
            FCHN = driver.find_element_by_xpath(LFCH).text
            print ("YA CON MODIFICATORIO", FCHN)
            driver.find_element_by_xpath("//table[@id='contentDS']/tbody[1]/tr[3]/td[6]/select[1]").send_keys("Eliminar" + Keys.ENTER)
            Alert(driver).accept()
            time.sleep(2)
            Alert(driver).accept()
            time.sleep(2)
        except NoSuchElementException:
            if fecha is None and monto is None:
                time.sleep(0.01)
            else:
                if juridico == contrato:
                    # SE DA CLICK AL MODIFICATORIO PARA ELABORAR NUEVO MODIFICATORIO
                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'icon')))
                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'typeFile'))).send_keys("Modificatorio" + Keys.ENTER)
                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, 'UltimoModificatorio')))
                    if fecha is not None:
                        # INGRESA LOS DATOS DEL MODIFICATORIO O REALIZAR
                        MES = ['Nel', 'Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun', 'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']
                        try: 
                            DIG1 = int(fecha.strftime("%m"))
                            DIA = str(int(fecha.strftime("%d")))
                            ANNO = fecha.strftime("%Y")
                        except:
                            # SI LA FECHA VIENE EN FORMATO TEXTO SE CAMBIA A FORMATO FECHA
                            fecha = datetime.datetime.strptime(fecha, '%d/%m/%Y')
                            DIG1 = int(fecha.strftime("%m"))
                            DIA = str(int(fecha.strftime("%d")))
                            ANNO = fecha.strftime("%Y")
                        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'isfecha'))).click()
                        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, 'ui-datepicker-trigger'))).click()
                        time.sleep(0.5)
                        MONTH = Select(driver.find_element_by_class_name('ui-datepicker-month'))
                        MONTH.select_by_visible_text(MES[DIG1])
                        YEAR = Select(driver.find_element_by_class_name('ui-datepicker-year'))
                        YEAR.select_by_visible_text(ANNO)
                        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.LINK_TEXT, DIA))).click()
                    if fcha2 is not None:
                            try:
                                fcha2 = datetime.datetime.strptime(fcha2, "%d/%m/%Y")
                            except:
                                time.sleep(0.01)
                            fchaucm = datetime.datetime.strptime(WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.XPATH, "//table[@id='convenioModificatorio']/tbody[1]/tr[7]/td[2]/span[1]"))).text, "%d/%m/%Y")
                            print(fcha2, fchaucm, fcha2 > fchaucm)
                            MESS2 = ['Nel', '01', '02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12']
                            DIG12 = int(fcha2.strftime("%m"))
                            MES2 = MESS2 [DIG12]
                            DIA2 = str(int(fcha2.strftime("%d")))
                            ANNO2 = fcha2.strftime("%Y")
                    if monto is not None:
                        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'ismonto'))).click()
                        importe = str(monto)
                        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'montoIG'))).send_keys(importe)
                        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'montoIG'))).send_keys(Keys.TAB)
                    print (contrato, "modificado por:", monto, "y", DIA+"/"+str(fecha.strftime("%m"))+"/"+str(ANNO), "firmado el:",fcha2)
                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'firmaCM'))).send_keys(str(int(DIA2) - 2)+"/"+str(MES2)+"/"+str(ANNO2))
                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'sig'))).click()
                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'sig'))).click()
                    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'acc'))).click()
                    time.sleep(3)


def finiquito (driver, pagina, juridico):
    driver.get(pagina)
    contrato = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'titleContrato'))).text
    try:
        HREF = "a[href*='/juridico/Contratos_Juridico/" + contrato + "/" + contrato + "_CF.aspx']"
        FINIQ = driver.find_element_by_css_selector(HREF).text
        print ("YA FINIQUITADO", FINIQ)
    except NoSuchElementException:
        if juridico == contrato:    
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'icon')))
            WebDriverWait(driver, 12).until(EC.element_to_be_clickable((By.ID, 'typeFile'))).send_keys("Finquito"+Keys.ENTER)
            time.sleep(1)
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'sig'))).click()
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'acc'))).click()
            time.sleep(3)
            driver.refresh()
            print(juridico, "Finiquitado con éxito.")


def eliminar_todo (driver, contraseña, iid, juridico):
    pagina = "http://fprado@javer.com.mx:" + contraseña + "@http://portal.javer.net/juridico/Paginas/Biblioteca.aspx?ContratoID=" + iid
    driver.get(pagina)
    time.sleep(1)
    contrato = WebDriverWait(driver, 25).until(EC.presence_of_element_located((By.XPATH, "//td[@id='titleContrato']//h2[1]"))).text
    if juridico == contrato:
        WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, 'icon')))
        WebDriverWait(driver, 12).until(EC.element_to_be_clickable((By.ID, 'btnEliminarContrato'))).click()
        time.sleep(2)
        alert_obj = driver.switch_to.alert
        alert_obj.accept()
        time.sleep(2)
        alert_obj.accept()
        driver.refresh()


def buscar_fecha (driver):
    FF = openpyxl.load_workbook('C:/Users/fprado/REPORTES/BDD/Z.xlsx')
    SFF = FF['F']
    e_d = openpyxl.load_workbook('C:/Users/fprado/REPORTES/BDD/SUM-JAV.xlsx')
    AUT = e_d['FeFi']
    A = []
    B = [] 
    C = []
    NUM = []
    N = 0
    for cell in AUT['A']:
        if cell.value != 0:
            NUM.append(cell.row)
            CONST = str(NUM[N])
            A.append(AUT['A' + CONST].value)
            B.append(AUT['B' + CONST].value)
            C.append(AUT['C' + CONST].value)
            N += 1
    print(N)
    acc_fir_oracle(driver)
    ir_en_fav(driver, 'Sistema de reportes')
    BUSPROY = "/html/body/form/span[2]/div/div[3]/div[1]/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr[1]/td[3]/span/a/img"
    BUSFREN = "/html/body/form/span[2]/div/div[3]/div[1]/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr[2]/td[3]/span/a/img"
    BUSCONJ = "/html/body/form/span[2]/div/div[3]/div[1]/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr[3]/td[3]/span/a/img"
    BUSPROV = "/html/body/form/span[2]/div/div[3]/div[1]/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr[4]/td[3]/span/a/img"
    BUSORDE = "/html/body/form/span[2]/div/div[3]/div[1]/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[2]/table/tbody/tr/td[2]/table/tbody/tr[5]/td[3]/span/a/img"
    #CANCELAR = "(//div[@class='x79']//table)[2]/tbody[1]/tr[1]/td[2]/button[1]"
    for x in range(N):
        ir_en_rg(driver, 'Reporte de estimacion para contratista')
        count = 0
        while True:
            if count == 5:
                break
            try:
                PROY, FRENTE, ETAPA = CONJUNTO(B[x]) 
                WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, BUSPROY))).click()
                frame_click_go(driver, PROY)
                WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Nombre del frente"]'))).click()
                WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, BUSFREN))).click()
                frame_click_go(driver, FRENTE)
                WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Nombre del conjunto"]'))).click()
                WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, BUSCONJ))).click()
                frame_click_go(driver, B[x])
                WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Nombre del proveedor"]'))).click()
                WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, BUSPROV))).click()
                frame_click_go(driver, C[x])        
                WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Orden de compra de la estimacion"]'))).click()
                WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, BUSORDE))).click()
                break
            except:
                count += 1
                driver.switch_to.window(driver.window_handles[0])
                print(count)
                continue
        count = 0
        while True:
            if count == 5:
                print("falló")
                break
            try:
                print(count, end="|")
                driver.switch_to.window(driver.window_handles[1])
                driver.switch_to.frame(0)
                break
            except:
                time.sleep(1)
                count += 1
                continue
        try:
            WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, "(//form[@id='_LOVResFrm']//table)[3]/tbody[1]/tr[1]/td[1]/div[1]/div[1]/input[1]"))).send_keys("%")
            WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, "(//form[@id='_LOVResFrm']//table)[3]/tbody[1]/tr[1]/td[1]/div[1]/div[1]/button[1]"))).click()
            FECHAF = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.ID, "N1:displayColumn4:0"))).text
            FECHAFM = FECHAF[8:10] + '/' + FECHAF[5:7] + '/' + FECHAF[0:4]
            CELDC = 'A' + str(x + 1)
            CELD = 'B' + str(x + 1)
            SFF[CELDC] = A[x]
            SFF[CELD] = FECHAFM
            print (CELD, A[x], FECHAFM)
        except:
            driver.close()
            time.sleep(1) 
            driver.switch_to.window(driver.window_handles[0])
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Logo Oracle"]')))
            driver.switch_to.window(driver.window_handles[0])
            WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[id="ButtonCancel"]'))).click()
            continue
        driver.close()
        time.sleep(1)
        driver.switch_to.window(driver.window_handles[0])
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[title="Logo Oracle"]')))
        driver.switch_to.window(driver.window_handles[0])
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[id="ButtonCancel"]'))).click()
        FF.save('C:/Users/fprado/REPORTES/BDD/Z.xlsx')


# PROGRAMAS -------------------------------------------------------------------------------------------------------------------------------------------------------------


def LISTRECC(driver):
    BDD_LIS = 'C:/Users/fprado/REPORTES/BDD/Z.xlsx'
    EL = [[], []]
    N = 0
    A = pd.read_excel(BDD_LIS, sheet_name="ORG", header=None)[0].drop_duplicates()
    print(A)
    acc_fir_oracle(driver)
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//table[@class='x6w']/tbody[1]/tr[1]/td[3]/a[1]"))).click()
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "(//li[@class='rootmenu']//a)[1]"))).click()
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//table[@id='respList']/tbody[1]/tr[1]/td[1]/ul[1]/li[1]/ul[1]/li[3]/a[1]"))).click()
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//li[@id='Frentes']//a[1]"))).click()
    Z = 0
    for x in A:
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "OrganizationLOV"))).send_keys(x)
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "span#OrganizationLOV__xc_0>a>img"))).click()
        time.sleep(1)
        driver.switch_to.window(driver.window_handles[1])
        time.sleep(1)
        driver.switch_to.frame(0)
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "N1:N8:0"))).click()
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "(//div[@class='x79']//table)[2]/tbody[1]/tr[1]/td[4]/button[1]"))).click()
        time.sleep(1)
        driver.switch_to.window(driver.window_handles[0])
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "Search"))).click()
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "FrontsTable:ResultsDisplayed:0"))).send_keys("100")
        time.sleep(1)
        Y = 0
        while True:
            try:
                BUTT = "FrontsTable:Code:" + str(Y)
                CODIG = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, BUTT))).text
                print(x, " - ", CODIG)
                EL[0].append(x)
                EL[1].append(CODIG)
                Y += 1
                continue
            except:
                break
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "OrganizationLOV"))).clear()
    pd.DataFrame(list(zip(EL[0], EL[1]))).to_excel(BDD_LIS, sheet_name="ORG", header=False, index=False)
    print("LISTONES PAPS")


def STATCONJ(driver):
    RUT_BAS = "C:\\Users\\fprado\\REPORTES\\BDD\\Z.xlsx"
    LIS_ORG = pd.read_excel(RUT_BAS, sheet_name="ORG", header=None, dtype=object)
    LIS_FRE = pd.Series(LIS_ORG[1], index=LIS_ORG[0])
    print (LIS_FRE)
    # ORGS = LIS_ORG.index.drop_duplicates()
    # for O in ORGS:
        # print(O)
        # print(LIS_ORG.loc[O][1])
        # for F in LIS_ORG.loc[O][0]:
            # print(F)
    '''
    acc_fir_oracle(driver)
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//table[@class='x6w']/tbody[1]/tr[1]/td[3]/a[1]"))).click()
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "(//li[@class='rootmenu']//a)[1]"))).click()
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//table[@id='respList']/tbody[1]/tr[1]/td[1]/ul[1]/li[1]/ul[1]/li[3]/a[1]"))).click()
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//li[@id='Frentes']//a[1]"))).click()
    '''
    # for O in ORGS:
        # print(LIS_ORG.groupby(level=0).mean())
    '''
    for c, x, y in zip(LIS_ORG.index, LIS_ORG[0], LIS_ORG[1]):
        print(c, x, y)
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "OrganizationLOV"))).send_keys(x)
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "span#OrganizationLOV__xc_0>a>img"))).click()
        time.sleep(1)
        driver.switch_to.window(driver.window_handles[1])
        time.sleep(1)
        driver.switch_to.frame(0)
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "N1:N8:0"))).click()
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "(//div[@class='x79']//table)[2]/tbody[1]/tr[1]/td[4]/button[1]"))).click()
        time.sleep(1)
        driver.switch_to.window(driver.window_handles[0])
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "Search"))).click()
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "FrontsTable:ResultsDisplayed:0"))).send_keys("100")
        time.sleep(1)
    '''

'''
    BKK = openpyxl.load_workbook('C:/Users/fprado/REPORTES/BDD/SUM-JAV.xlsx')
    AUT = BKK['OGG']
    SAV = BKK['ORG']
    A, EL, NUM = [], [], []
    N = 0
    for cell in AUT['A']:
        if cell.value != 0:
            NUM.append(cell.row)
            CONST = str(NUM[N])
            if AUT['A' + CONST].value != None:
                A.append(AUT['A' + CONST].value)
            N += 1
    print(A)
    acc_fir_oracle(driver)
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//table[@class='x6w']/tbody[1]/tr[1]/td[3]/a[1]"))).click()
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "(//li[@class='rootmenu']//a)[1]"))).click()
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//table[@id='respList']/tbody[1]/tr[1]/td[1]/ul[1]/li[1]/ul[1]/li[3]/a[1]"))).click()
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//li[@id='Frentes']//a[1]"))).click()
    Z = 0
    for x in A:
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "OrganizationLOV"))).send_keys(x)
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "span#OrganizationLOV__xc_0>a>img"))).click()
        time.sleep(1)
        driver.switch_to.window(driver.window_handles[1])
        time.sleep(1)
        driver.switch_to.frame(0)
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "N1:N8:0"))).click()
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "(//div[@class='x79']//table)[2]/tbody[1]/tr[1]/td[4]/button[1]"))).click()
        time.sleep(1)
        driver.switch_to.window(driver.window_handles[0])
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "Search"))).click()
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "FrontsTable:ResultsDisplayed:0"))).send_keys("100")
        time.sleep(1)
        Y = 0
        while True:
            try:
                BUTT = "FrontsTable:Code:" + str(Y)
                CODIG = x + WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, BUTT))).text
                print(CODIG)
                EL.append(CODIG)
                SAV['A' + str(Z + 1)] = CODIG
                Z += 1
                Y += 1
                continue
            except:
                break
        WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "OrganizationLOV"))).clear()
    BKK.save('C:/Users/fprado/REPORTES/BDD/SUM-JAV.xlsx')
    print("LISTONES PAPS")
    '''


def RECC(driver):
    # Descarga los estados de cuenta ingresados en SUM-JAV
    SJ = openpyxl.load_workbook('C:/Users/fprado/REPORTES/BDD/SUM-JAV.xlsx')
    AUT = SJ['ORG']
    A, B, C, NUM = [], [], [], []
    N = 0
    for cell in AUT['A']:
        if cell.value != 0 and cell.value != None:
            NUM.append(cell.row)
            CONST = str(NUM[N])
            A.append(AUT['A' + CONST].value[:3])
            B.append(AUT['A' + CONST].value[3:])
            N += 1
    acc_fir_oracle(driver)
    for x in range(N):
        print(A[x], "-",B[x])
        try:
            C.append(datos_recc(driver, A[x], B[x])[-8:])
        except:
            C.append(datos_recc(driver, A[x], B[x])[-8:])
    time.sleep(20)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "Refresh"))).click()
    descargarecc(driver, C)
    time.sleep(10)
    moveRECC('A')
    print("ÉXITO")


def RECC2 (driver):
    SJ = openpyxl.load_workbook('C:/Users/fprado/REPORTES/BDD/SUM-JAV.xlsx')
    AUT = SJ['ORG']
    A, B, C, NUM = [], [], [], []
    N = 0
    for cell in AUT['A']:
        if cell.value != 0 and cell.value != None:
            NUM.append(cell.row)
            CONST = str(NUM[N])
            A.append(AUT['A' + CONST].value[:3])
            B.append(AUT['A' + CONST].value[3:])
            N += 1
    acc_fir_oracle(driver)
    for Z in range(N):
        while True:
            # try:
            C.append(datos_recc2(driver, A[Z], B[Z])[-8:])
            break
            # except:
                # continue
    time.sleep(20)
    WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.ID, "Refresh"))).click()
    dfcsv = pd.DataFrame(C, columns=["Requests"])
    dfcsv.to_csv("C:/Users/fprado/REPORTES/BDD/Requests.csv", index=False)
    lista = pd.read_csv("C:/Users/fprado/REPORTES/BDD/Requests.csv")
    listaa = lista["Requests"].tolist()
    descargarecc(driver, listaa)
    time.sleep(10)
    print("ÉXITO")


def ESTIMACIONES(FCHI, FF):
    driver = create_driver(headless = False)
    # Descarga las esitmaciones guardadas en el libro de excel estimaciones
    PAT = "C:/Users/fprado/REPORTES/BDD/ESTIMACIONES.xlsx"
    if FCHI == "":
        FI = lastsaturday()
    else:
        FI = FCHI
    FECHA = FI
    acc_fir_oracle(driver)
    ir_en_fav(driver, 'Sistema de reportes')
    # return N, B, C, D, E, F
    N, C, PR, SOBC, SOBD, SOBE, SOBF = EXC('AUT', PAT)
    AX = 0
    while True:
        AY = AX + 1
        if AX == len(C):
            break
        while True:
            if AY == len(C):
                break
            if C[AX] + PR[AX] == C[AY] + PR[AY]:
                print("se quitó este", C[AX] + " | " + PR[AX], C[AY] + " | " + PR[AY])
                N -= 1
                C.pop(AY)
                PR.pop(AY)
                continue
            AY += 1
        AX += 1
    AX = 0
    while True:
        AY = AX + 1
        if AX == len(C):
            break
        while True:
            if AY == len(C):
                break
            if C[AX][7] == "C" or C[AX][7] == "E" or C[AX][7] == "P":
                if C[AX][:6] + PR[AX] == C[AY][:6] + PR[AY]:
                    print("se quitó este", C[AX][:6] + " | " + PR[AX], C[AY][:6] + " | " + PR[AY])
                    N -= 1
                    C.pop(AY)
                    PR.pop(AY)
                    continue
            else:
                pass
            AY += 1
        AX += 1    
    for x in range(N):
        COUNT = 1
        while True:
            P, F, B = CONJUNTO(C[x])
            print(x + 1, "/", N, ":", COUNT, ":",P, "|", F, "|", B, "|", C[x], " ", PR[x], " ", sep="", end="|")
            if COUNT == 3:
                print("Falló:", x + 1)
                break
            if B == 'C03' or B == 'P05' or B == 'E04':
                try:
                    ir_en_rg(driver, 'Reporte de estimaciones para contratista detalle viviendas sembrado')
                    print("RG", end="| ")
                    datos_repcdvs(driver, P, F, C[x], PR[x], FI, FF)
                    objetivo(driver, 'PDF', 'Reporte de estimaciones para contratista detalle viviendas sembrado')
                    COUNT = 1
                    print("OK")
                    break
                except TimeoutException:
                    COUNT += 1
                    print("  ")
                    acc_fir_oracle(driver)
                    ir_en_fav(driver, 'Sistema de reportes')
                    continue
                except:
                    print ("- Falló")
                    break
            else:
                EC = C[x]
                EPR = PR[x]
                OC = []
                ir_en_rg(driver, 'Reporte de estimacion para contratista')
                print("     EE", end="| ")
                OC = ecuentra_estimaciones(driver, P, F, C[x], PR[x], FECHA)
                print("son: ", len(OC))
                COUNT = 0
                VAR = 0
                while True:
                    if COUNT == 3:
                        VAR += 1
                        break
                    if VAR == len(OC):
                        break
                    try:
                        print ("    1", EC, " ", EPR, OC[VAR], end=" | ")
                        ir_en_rg(driver, 'Reporte de estimacion para contratista')
                        print("RG", end="| ")
                        datos_repepc(driver, P, F, EC, EPR, FECHA, OC[VAR])
                        objetivo(driver, 'PDF', 'Reporte de estimacion para contratista')
                        print("OK")
                        VAR += 1
                        continue
                    except TimeoutException:
                        COUNT += 1
                        print("  ")
                        acc_fir_oracle(driver)
                        ir_en_fav(driver, 'Sistema de reportes')
                        continue
                    except:
                        COUNT += 1
                        print ("- Falló")
                        break
                break
    print("C'est fini mon amie")
    time.sleep(8)
    driver.quit()


def CUADRICULAS(driver, ETAPA):
    # Descarga las cuadriculas excluyendo la etapa o etapas señaladas, guardadas en el libro de excel ESTIMACIONES
    path = 'C:/Users/fprado/REPORTES/BDD/ESTIMACIONES.xlsx'
    acc_fir_oracle(driver)
    ir_en_fav(driver, 'Sistema de reportes')
    N, C, PR, SOBC, SOBD, SOBE, SOBF = EXC('AUT', path)
    AX = 0
    while True:
        AY = AX + 1
        if AX == len(C):
            break
        while True:
            if AY == len(C):
                break
            if PR == None:
                break
            if C[AX] + PR[AX] == C[AY] + PR[AY]:
                print("se quitó este", C[AX] + " | " + PR[AX], C[AY] + " | " + PR[AY])
                N -= 1
                C.pop(AY)
                PR.pop(AY)
                continue
            AY += 1
        AX += 1
    for x in range(N):
        COUNT = 0
        while True:
            P, F, B = CONJUNTO(C[x])
            COUNT += 1
            if COUNT == 5:
                break
            if B[0] != ETAPA:
                print(N - x, "|", B[0],"|", COUNT, "|", C[x], "-", PR[x])
                try:
                    ir_en_rg(driver, 'Reporte de matriz de control de contratos')
                    datos_rmcc(driver, P, F, C[x], PR[x])
                    objetivo(driver, 'XLS', 'Reporte de matriz de control de contratos')
                    break
                except TimeoutException:
                    acc_fir_oracle(driver)
                    ir_en_fav(driver, 'Sistema de reportes')
                    N, C, PR, SOBC, SOBD, SOBE, SOBF = EXC('AUT', path)
                    continue
            else:
                print(N - x, "|", B[0])
                break
    time.sleep(5)
    print("Ces't fini")


def INSUMOS(driver):
    # Espera a ingresar actividades en contrato y descarta los insumos de la lista guradada en el libro de excel SUM-JAV, para elaborar contrato
    acceder_oracle(driver)
    go_contract(driver)
    BK = openpyxl.load_workbook('C:/Users/fprado/REPORTES/BDD/SUM-JAV.xlsx')
    AUT = BK['SUM']
    A, NUM = [], []
    N = 0
    COUNT = 0
    for cell in AUT['A']:
        if cell.value != 0 and cell.value is not None:
            NUM.append(cell.row)
            CONST = str(NUM[N])
            A.append(AUT['A' + CONST].value)
            N += 1
    print(N)
    WebDriverWait(driver, 9999).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ActivityType"]'))).click()
    ACTIVIDAD = Select(driver.find_element(By.XPATH, '//*[@id="ActivityType"]'))
    ACTIVIDAD.select_by_visible_text("Material")
    time.sleep(5)
    SIG = '/html/body/form/span[2]/div/div[3]/div/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[3]/table/tbody/tr[4]/td/table/tbody/tr/td/div/div/span[1]/table[1]/tbody/tr[1]/td/table/tbody/tr/td[5]/table/tbody/tr/td[9]/a/img'
    SIG2 = '/html/body/form/span[2]/div/div[3]/div/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[3]/table/tbody/tr[4]/td/table/tbody/tr/td/div/div/span[1]/table[1]/tbody/tr[1]/td/table/tbody/tr/td[5]/table/tbody/tr/td[9]/img'
    while True:
        for x in range(15):
            COD = []
            while True:
                X = str(x)
                Y = str(x + 2)
                ELEM = '//*[@id="ResourcesTable:ResourceCode:' + X + '"]'
                BOTT = '/html/body/form/span[2]/div/div[3]/div/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[3]/table/tbody/tr[4]/td/table/tbody/tr/td/div/div/span[1]/table[2]/tbody/tr[' + Y + ']/td[1]/input'
                try:
                    COD = (WebDriverWait(driver, 0.1).until(EC.element_to_be_clickable((By.XPATH, ELEM))).text)
                    time.sleep(0.1)
                    if COD in A:
                        while True:
                            try:
                                WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, BOTT))).click()
                                time.sleep(0.2)
                                COUNT += 1
                                A.remove(COD)
                                print(COUNT, "-", len(A), "|", COD)
                                break
                            except StaleElementReferenceException:
                                time.sleep(0.1)
                                continue
                            except TimeoutException:
                                time.sleep(0.1)
                                continue
                except StaleElementReferenceException:
                    time.sleep(0.1)
                    continue
                except TimeoutException:
                    time.sleep(0.1)
                    continue
                break
        print(A)
        try:
            STAT = WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, SIG2))).get_attribute('title')
        except:
            STAT = WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, SIG))).get_attribute('title')
        print(STAT)
        if STAT == 'Next functionality disabled' or len(A) == 0:
            print(STAT)
            break
        else:
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/span[2]/div/div[3]/div/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[3]/table/tbody/tr[4]/td/table/tbody/tr/td/div/div/span[1]/table[1]/tbody/tr[1]/td/table/tbody/tr/td[5]/table/tbody/tr/td[7]/a'))).click()
            time.sleep(0.2)
            continue
    print('TERMINÓ')


def MATRICES(driver):
    e_d = openpyxl.load_workbook('C:/Users/fprado/REPORTES/BDD/SUM-JAV.xlsx')
    AUT = e_d.get_sheet_by_name('MAT')
    A = []
    NUM = []
    N = 0
    for cell in AUT['A']:
        if cell.value != 0:
            NUM.append(cell.row)
            CONST = str(NUM[N])
            A.append(AUT['A' + CONST].value)
            N += 1
    acc_fir_oracle(driver)
    WebDriverWait(driver, 300).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="ResultsTable:item21:0"]')))
    time.sleep(4)
    SIG = '/html/body/form/span[2]/div/div[3]/div/div[2]/table[2]/tbody/tr[4]/td/table/tbody/tr/td/div/div[3]/span[1]/table[1]/tbody/tr/td/table/tbody/tr/td[3]/table/tbody/tr/td[9]/a/img'
    while True:
        for x in range(15):
            COD = []
            DES = []
            while True:
                X = str(x)
                ELEM = '//*[@id="ResultsTable:item2:' + X + '"]'
                BOTT = '//*[@id="ResultsTable:item21:' + X + '"]'
                DESC = '//*[@id="ResultsTable:item31:' + X + '"]'
                try:
                    COD = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, ELEM))).text
                    DES = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, DESC))).text
                    print(BOTT, ' - ', COD)
                    time.sleep(0.1)
                    if COD in A:
                        time.sleep(1)
                        WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, BOTT))).click()
                        print(COD, DES)
                    break
                except StaleElementReferenceException:
                    time.sleep(0.05)
                    continue
                except TimeoutException:
                    break
        try:
            STAT = WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, SIG))).get_attribute('title')
        except:
            'NO HAY'
        print(STAT)
        if STAT == 'Next functionality disabled':
            print(STAT)
            break
        else:
            try:
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, SIG))).click()
            except StaleElementReferenceException:
                time.sleep(0.5)
            continue
    print('TERMINÓ')
    time.sleep(5)
    driver.close()


def AUTOFIN(driver, PASS, mod):
    """FINIQUITA LA LISTA GUARDADA EN EXCEL
        driver = explorador a usar
        mod = MOD, FIN o DEL 
        DONDE LAS COLUMNAS: 
            A = CONTRATO
            B = PAGINA
            C = MONTO MODIFICATORIO
            D = FECHA MODIFICATORIO
    """
    e_d = openpyxl.load_workbook('C:/Users/fprado/REPORTES/BDD/SUM-JAV.xlsx')
    AUT = e_d.get_sheet_by_name('FIN')
    A, B, C, D, E, NUM = [], [], [], [], [], []
    N = 0
    for cell in AUT['A']:
        if cell.value != 0:
            NUM.append(cell.row)
            CONST = str(NUM[N])
            HI = AUT['B' + CONST].value
            HYV = HI[0:7] + "fprado:" + PASS + "@" + HI[7:]
            A.append(AUT['A' + CONST].value)
            B.append(HYV)
            C.append(AUT['C' + CONST].value)
            D.append(AUT['D' + CONST].value)
            E.append(AUT['E' + CONST].value)
            N += 1
    print(N)
    x = 0
    if "MOD" in mod:
        for x in range(len(A)):
            while True:
                print (x + 1, "de", N)
                try:
                    modificatorio(driver, B[x], C[x], D[x], E[x], A[x])
                    break
                except TimeoutException:
                    print ("FAIL", A[x])
                    A.remove(A[x])
                    B.remove(B[x])
                    C.remove(C[x])
                    D.remove(D[x])
                    E.remove(E[x])
                    break
    if "FIN" in mod:
        N = len(A)
        for x in range(len(A)):
            while True:
                print(x + 1, "de", N)
                try:
                    finiquito(driver, B[x], A[x])
                    break
                except TimeoutException:
                    print ("FAIL", A[x])
                    break
    elif "DEL" in mod:
        for x in range(N):
            while True:
                print(x)
                try:
                    eliminar_todo(driver, B[x], A[x])
                    print (A[x], "ELIMINADO")
                    break
                except TimeoutException:
                    print ("FAIL", A[x])
                    break


def PLANTILLA(headless=True):
    driver = create_driver(headless=headless)
    acc_fir_oracle(driver)
    try:
        ir_PLANTILLA(driver)
    except:
        time.sleep(1)
    WebDriverWait(driver, 200  ).until(EC.presence_of_element_located((By.XPATH, "/html/body/form/span[2]/div/div[3]/div[3]/div[2]/table/tbody/tr[2]/td/span[1]/table[1]/tbody/tr[1]/td/table/tbody/tr/td/a[1]")))
    # return N, B, C, D, E, F
    N, A, B, C, D, E, F = EXC('A', 'C:\\Users\\fprado\\REPORTES\\INSUMOS.xlsx')
    print(N)
    COUNT = 0
    for x in range(N):
        XPATT = "//input[@id='N3:Y:" + str(A[x]) + "']"
        ID0 = "N3:Unidad:" + str(A[x])
        IDD = "N3:Y1:" + str(A[x])
        #print(XPATT, ID0, IDD)
        WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, XPATT)))
        value = WebDriverWait(driver, 1).until(EC.element_to_be_clickable((By.XPATH, XPATT))).get_attribute('value')
        if value != "":
            print(B[x], " - ", value)
        else:
            print(B[x], " - ", D[x], E[x])
            WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, XPATT))).send_keys(D[x] + Keys.TAB)
            while True:
                try:
                    listo = WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.ID, ID0))).text
                    if listo == 'LOTE':
                        break
                except:
                    time.sleep(0.5)
            WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.ID, IDD))).send_keys('1' + Keys.TAB)
            time.sleep(0.5)
            COUNT += 1
            if COUNT == 20:
                print("Guardando")
                WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="guardar"]'))).click()
                time.sleep(0.5)
                WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, '/html/body/form/span[2]/div/table/tbody/tr/td/table/tbody/tr[2]/td[2]/div[1]/div/table/tbody/tr/td[3]/table/tbody/tr/td/h1')))
                COUNT = 0
    WebDriverWait(driver, 50).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="guardar"]'))).click()
    time.sleep(5)
    driver.close()


def get_page_and_print(self, page, filepath):
        # Get about:config
        self.driver.get('about:config')
        time.sleep(1)

        # Define Configurations
        script = """
        var prefs = Components.classes['@mozilla.org/preferences-service;1'].getService(Components.interfaces.nsIPrefBranch);
        prefs.setBoolPref('print.always_print_silent', true);
        prefs.setCharPref('print_printer', 'Print to File');
        prefs.setBoolPref('print.printer_Print_to_File.print_to_file', true);
        prefs.setCharPref('print.printer_Print_to_File.print_to_filename', '{}');
        prefs.setBoolPref('print.printer_Print_to_File.show_print_progress', true);
        """.format(filepath)

        # Set Configurations
        self.driver.execute_script(script)
        time.sleep(1)

        # Get site to print in pdf
        self.driver.get(page)
        time.sleep(2)
        self.driver.execute_script("window.print();")


def getpurchaseorders():
    EXFIL = "C:/Users/fprado/OneDrive - Servicios Administrativos Javer, S.A. DE C.V/Descargas/CARATULAS/EST-SEMANAL" + ".xlsx"
    borrar =[]
    # Bases para el DataFrame
    rows = []

    # Rutas
    filepath = "C:\\Users\\fprado\\OneDrive - Servicios Administrativos Javer, S.A. DE C.V\\Descargas"

    # Iteramos archivos en carpetas
    for root, dirs, files in os.walk(filepath):
        for filename in files:
            basename, extention = os.path.splitext(filename)

            # Abrimos pdfs
            if extention == '.pdf':
                fullpath = root + '\\' + filename
                open_pdf = pdfplumber.open(fullpath)

                # Extraemos texto de todas las páginas
                with open_pdf as pdf:
                    try:
                        pdf_text = ''
                        for page in pdf.pages:
                            pdf_text += page.extract_text() + "\n"
                    except TypeError:
                        continue

                # Verificamos que sea estimación
                es_estimacion = re.search(r'Reporte de', pdf_text)
                if es_estimacion:
                    appender = []
                    try:
                        new_pdf_text = re.search(r'Reporte.*?\n(.*?)Código', pdf_text, re.DOTALL).group(0)
                        # Obtenemos expresiones regulares
                        conjunto_a = re.search(r'E\d{2}-\d{2}-[A-Za-z]\d{2}-\d{2}-\d{3}', pdf_text)
                        conjunto_b = re.search(r'E\d{2}-\d{2}-[A-Za-z]\d{2}-\d{2}', pdf_text)
                        if conjunto_a:
                            conjunto = conjunto_a.group(0)
                        elif conjunto_b:
                            conjunto = conjunto_b.group(0)
                        descripcion = re.search(re.escape(conjunto) + r'(.*?)\n', pdf_text).group(1)
                        
                        # Contrato
                        contrato_a = re.search(r'Contrato (.*?) - ', pdf_text)
                        contrato_b = re.search(re.escape(descripcion) + r'\n(.*?) - ', pdf_text)
                        contrato_c = re.search(re.escape(descripcion) + r'\nConjunto\n.*?\n(.*?) - ', pdf_text)
                        # contrato_c = re.search()
                        if contrato_a:
                            contrato = contrato_a.group(1)
                            print("A la primera: ", contrato)
                        elif contrato_b:
                            contrato = contrato_b.group(1)
                            print("A la segunda: ", contrato)
                        elif contrato_c:
                            contrato = contrato_c.group(1)
                            print("A la tercera: ", contrato)
                        else:
                            print("Este falló: ", filename, descripcion)
                            print(new_pdf_text)
                        
                        # Proveedor
                        proveedor = re.search(r'Proveedor (.*?) Empresa', pdf_text).group(1)

                        # Orden de compra
                        orden_compra = re.search(r'Compra (\d+)\nFecha', pdf_text)
                        if orden_compra:
                            orden_compra = orden_compra.group(1)
                        else:
                            orden_compra = "Sin_OC"

                        # Importes
                        total_estimar = re.search(r'Total de la estimación (.*?)\n', pdf_text).group(1)
                        total_estimar = re.search(r'(^.*?) ', total_estimar).group(1)
                        total_pagar = re.search(r'Total a pagar (.*?)\n', pdf_text).group(1)
                        total_pagar = re.search(r' .*? (.*?$)', total_pagar).group(1)

                        # Guardamos valores
                        appender.append(proveedor)
                        appender.append(contrato)
                        appender.append(conjunto)
                        appender.append(orden_compra)
                        appender.append(total_estimar)
                        appender.append(total_pagar)
                        rows.append(appender)

                        # Renombramos el archivo en caso de no estar en caratulas
                        if basename.upper().startswith('XXM'):
                            doc_name = filepath + "\\CARATULAS\\" + proveedor + ' - ' + conjunto + ' - ' + orden_compra + '.pdf'
                            print(doc_name)
                            open_pdf.close()
                            try:
                                os.rename(fullpath, doc_name)
                            except FileExistsError as E:
                                print("error file exist", E)
                                borrar.append(fullpath)
                        if orden_compra == "Sin_OC" and total_pagar == "0.00" and total_estimar == "0.00":
                            borrar.append(fullpath)

                    except Exception as e:
                        print(e)
                        print(pdf_text)
                        time.sleep(6)

                else:
                    print(filename)

    columns = ['Proveedor', 'Contrato', 'Conjunto', 'Orden de Compra', 'Total a Estimar', 'Total a Pagar']
    df = pd.DataFrame(rows, columns=columns)
    df.to_excel(EXFIL, index=False)


    time.sleep(1)
    for borra in borrar:
        os.remove(borra)

def splitPDF():
    DIVIDIR = []
        # Ruta completa a la carpeta 'CARATULAS'
    caratulas_path = 'C:\\Users\\fprado\\OneDrive - Servicios Administrativos Javer, S.A. DE C.V\\Descargas\\CARATULAS'

    # Verificar si la carpeta 'CARATULAS' no existe y crearla si es necesario
    if not os.path.exists(caratulas_path):
        os.makedirs(caratulas_path)
        print(f'Se ha creado la carpeta {caratulas_path}')
    else:
        print(f'La carpeta {caratulas_path} ya existe')
    filepath = "C:\\Users\\fprado\\OneDrive - Servicios Administrativos Javer, S.A. DE C.V\\Descargas\\"
    for root, dirs, files in os.walk(filepath):
        for filename in files:
            print(filename)
            basename, extention = os.path.splitext(filename)
            try:
                if extention == '.pdf' and basename[:6] == 'XXMCAN':
                    AA, AB = [], []
                    count = 0
                    fullpath = os.path.join(filepath, filename)
                    open_pdf = pdfplumber.open(fullpath)
                    with open_pdf as pdf:
                        if len(pdf.pages) < 2:
                            pass
                        for page in pdf.pages:
                            pdf_text = page.extract_text()
                            if re.search(r'Reporte de estimacion', pdf_text) != None:
                                count += 1
                                AA.append(page)
                                AB.append(pdf.pages.index(page) + 1)
                                print(count)
                        if count > 1:
                            DATOS = {
                                'File': filename,
                                'Pages': AB,
                                'LastPage': len(pdf.pages),
                            }
                            DIVIDIR.append(DATOS)
            except FileNotFoundError:
                print(f"No existe {open_pdf}")
            except PDFSyntaxError as e:
                print(f"Error {e}")
                os.remove(fullpath)
    for ELEMENT in DIVIDIR:
        ARCHIVOI = filepath + ELEMENT['File']
        for PGINI in ELEMENT['Pages']:
            OUTPUT = PdfWriter()
            if ELEMENT['Pages'].index(PGINI) == len(ELEMENT['Pages']) - 1:
                PGULT = ELEMENT['LastPage']
            else:
                PGULT = ELEMENT['Pages'][ELEMENT['Pages'].index(PGINI) + 1] - 1
            ARCHIVOF = filepath + ELEMENT['File'][:-4] + ' ' + str(ELEMENT['Pages'].index(PGINI) + 1) + '.pdf'
            for X in range (PGINI, PGULT + 1):
                INPUT = PdfReader(ARCHIVOI)
                OUTPUT.add_page(INPUT.pages[X - 1])
            with open(ARCHIVOF, 'wb') as fh:
                OUTPUT.write(fh)
    for ELEMENT in DIVIDIR:
        try:
            ARCHIVOI = filepath + ELEMENT['File']
            os.remove(ARCHIVOI)
        except:
            continue

def oldpdf():
    x = "C:\\Users\\fprado\\OneDrive - Servicios Administrativos Javer, S.A. DE C.V\\Descargas\\"

    # Obtener la fecha de inicio y fin de la semana pasada
    hoy = datetime.datetime.now()
    inicio_semana_pasada = hoy - datetime.timedelta(days=hoy.weekday() + 7)
    fin_semana_pasada = inicio_semana_pasada + datetime.timedelta(days=6)

    for root, dirs, files in os.walk(x):
        for filename in files:
            ruta_archivo = os.path.join(root, filename)
            fecha_creacion = datetime.datetime.fromtimestamp(os.path.getctime(ruta_archivo))

            # Verificar si la fecha de creación está en la semana pasada
            if inicio_semana_pasada <= fecha_creacion <= fin_semana_pasada:
                try:
                    os.remove(ruta_archivo)
                except PermissionError as e:
                    print(e)

    print("Exito, removiendo archivos viejos")

def renamePDF():
    # Ruta completa a la carpeta 'CARATULAS'
    caratulas_path = 'C:\\Users\\fprado\\OneDrive - Servicios Administrativos Javer, S.A. DE C.V\\Descargas\\CARATULAS'

    # Verificar si la carpeta 'CARATULAS' no existe y crearla si es necesario
    if not os.path.exists(caratulas_path):
        os.makedirs(caratulas_path)
        print(f'Se ha creado la carpeta {caratulas_path}')
    else:
        print(f'La carpeta {caratulas_path} ya existe')
    borrar = []
    filepath = "C:\\Users\\fprado\\OneDrive - Servicios Administrativos Javer, S.A. DE C.V\\Descargas"
    for root, dirs, files in os.walk(filepath):
        for filename in files:
            try:
                basename, extention = os.path.splitext(filename)
                if extention == '.pdf' and basename[:3] == 'XXM':
                    print(basename)
                
                    # Extraemos texto
                    fullpath = root + '\\' + basename + extention
                    with pdfplumber.open(fullpath) as open_pdf:

                        open_pdf = pdfplumber.open(fullpath)
                        page_obj = open_pdf.pages[0]
                        pdf_text = page_obj.extract_text()

                    # Confirmación de que es un reporte
                    confirmacion = re.search(r'Reporte de', pdf_text)
                    if confirmacion is not None:

                        print(pdf_text)
                        pdf_prov = re.search(r'Proveedor (.*?) Empresa', pdf_text).group(1)

                        # Extraemos el conjunto y aseguramos esté bien escrito
                        pdf_conj = re.search(r'E\d{2}-\d{2}-[A-Za-z]\d{2}-\d{2}-\d{3}', pdf_text)
                        pdf_conj_c = re.search(r'E\d{2}-\d{2}-[A-Za-z]\d{2}-\d{2}', pdf_text)
                        if pdf_conj:
                            pdf_conj = pdf_conj.group(0)
                        elif pdf_conj_c:
                            pdf_conj = pdf_conj_c.group(0)

                        FRAC, FREN, ETAP = CONJUNTO(pdf_conj)
                        SEMM = datetime.date.today().isocalendar()[1]
                        
                        # Se busca la coincidencia de pdf
                        pdf_search = re.search(r'Compra (.*?)\nFecha', pdf_text)
                        if pdf_search is not None:
                            pdf_ult = pdf_search.group(1)
                        else:
                            pdf_ult = "sin OC"

                        DOCNAM = filepath + "\\CARATULAS\\" + pdf_prov + ' - ' + pdf_conj + ' - ' + pdf_ult + '.pdf'
                        print(pdf_prov, pdf_conj, pdf_ult)
                        open_pdf.close()
                        try:
                            os.rename(fullpath, DOCNAM)
                        except FileExistsError:
                            borrar.append(fullpath)

                    else:
                        borrar.append(fullpath)
                        # os.remove(fullpath)

            except TimeoutException:
                continue
    for borra in borrar:
        os.remove(borra)


def getfilenames():
    CM, CF, FALTAN = [], [], []
    filepath = tkinter.filedialog.askdirectory(title='Select File Folder') 
    for root, dirs, files in os.walk(filepath):
        for filename in files:
            try:
                basename, extention = os.path.splitext(filename)
                #print(basename)
                TYPE = "C" + re.search(r'_C(.*)', basename).group(1)
                CONTRACT = re.search(r' - (.*)_C', basename).group(1)
                if TYPE[:2] == 'CF':
                    CF.append(CONTRACT)
                elif TYPE[:2] == 'CM':
                    if "_" in CONTRACT:
                        CONTRACT = re.search(r'(.*)_', CONTRACT).group(1)
                    CM.append(CONTRACT)
            except:
                continue
    for x in CM:
        if x not in CF:
            FALTAN.append(x)
    for x in FALTAN:
        print(x)
    if not filepath =='':
        tkinter.messagebox.showinfo(title='', message="Files list")


def getcontractnames():
    CM, CF, FALTAN = [], [], []
    filepath = tkinter.filedialog.askdirectory(title='Select File Folder') 
    for root, dirs, files in os.walk(filepath):
        for filename in files:
            try:
                basename, extention = os.path.splitext(filename)
                #print(basename)
                TYPE = "C" + re.search(r'_C(.*)', basename).group(1)
                CONTRACT = re.search(r' - (.*)_C', basename).group(1)
                if TYPE[:2] == 'CF':
                    CF.append(CONTRACT)
                elif TYPE[:2] == 'CM':
                    if "_" in CONTRACT:
                        CONTRACT = re.search(r'(.*)_', CONTRACT).group(1)
                    CM.append(CONTRACT)
            except:
                continue
    for x in CF:
        print(x)
    if not filepath =='':
        tkinter.messagebox.showinfo(title='', message="Files list")


def getcontractnames2():
    CF, FALTAN = [], []
    filepath = tkinter.filedialog.askdirectory(title='Select File Folder') 
    for root, dirs, files in os.walk(filepath):
        for filename in files:
            try:
                basename, extention = os.path.splitext(filename)
                #print(basename)
                CONTRAC = re.search(r' - (.*)\Z', basename).group(1)
                CONTRACT = re.search(r' - (.*)\Z', CONTRAC).group(1)
                CF.append(CONTRACT)
            except:
                continue
    for x in CF:
        print(x)
    if not filepath =='':
        tkinter.messagebox.showinfo(title='', message="Files list")


def CUANTOS_MODIF(DRIVER):
    RUTA2 = "C:/Users/fprado/REPORTES/BDD/SUM-JAV.xlsx"
    RUTA = "C:/Users/fprado/REPORTES/EJJ.xlsm"
    e_d = openpyxl.load_workbook(RUTA)
    e_d2 = openpyxl.load_workbook(RUTA2)
    MOD = e_d['M']
    URL = e_d2['URL']
    NUM, NUM2, E, F, C, CC, NA, NB, CN= [], [], [], [], [], [], [], [], []
    N = 0
    for cell in MOD['A']:
        if cell.value != 0:
            NUM.append(cell.row)
            CONST = str(NUM[N])
            E.append(MOD['E' + CONST].value)
            F.append(MOD['F' + CONST].value)           
            N += 1
    N = 0
    for cell in URL['A']:
        if cell.value != 0:
            NUM.append(cell.row)
            CONST = str(NUM[N])
            C.append(URL['A' + CONST].value)
            CC.append(URL['B' + CONST].value)
            N += 1
    for y in range(len(C)):
        NA.append(C[y])
        NB.append(C[y] + "_CF")
        CN.append(CC[y])
        FCOUNT = []
        for x in range(len(F)):
            if F[x] == C[y]:
                NA.append(F[x])
                NB.append(E[x])
                CN.append(CC[y])
    for w in range(len(NB)):
        print(len(NB), "-", x, NB[w], CN[w])
        DRIVER.get("http://fprado@javer.com.mx:" + JU_PASSWORD + "@portal.javer.net/juridico/Contratos_Juridico/" + NA[w] + "/" + NB[w] + ".aspx")
        try:
            WebDriverWait(DRIVER, 2).until(EC.element_to_be_clickable((By.XPATH, "//img[@src='/_LAYOUTS/15/JAVERcontratos/img/LogoJaver_2020.png']")))
        except TimeoutException:
            WebDriverWait(DRIVER, 2).until(EC.element_to_be_clickable((By.XPATH, "//img[@src='/_LAYOUTS/15/JAVERcontratos/img/logo_javer.png']")))
        except TimeoutException:
            print(NB[w], CN[w], " -NO EXISTÉ EN SISTEMA")
            break
        pyautogui.moveTo(683, 384)
        pyautogui.click()
        pyautogui.click()
        pyautogui.hotkey('ctrl', 'p')
        time.sleep(1)
        pyautogui.press('G')
        time.sleep(0.1)
        pyautogui.press('G')
        time.sleep(0.8)
        pyautogui.press('enter')
        time.sleep(1)
        pyautogui.write(CN[w] + " - " + NB[w], interval=0.001)
        time.sleep(0.25)
        pyautogui.press('enter')
        time.sleep(0.40)
        pyautogui.press('enter')
        time.sleep(0.40)
        pyautogui.press('enter')
        time.sleep(0.45)
        pyautogui.press('enter')


def JURIDICO_DOWNLOADER(DRIVER):
    # Rutas de archivos
    NO_EXISTE_SISTEMA = []
    RUTA_POR_FINIQUITAR = "C:/Users/fprado/REPORTES/BDD/SUM-JAV.xlsx"
    RUTA_BUSCA_MODIFICATORIOS = "C:/Users/fprado/REPORTES/EJJ.xlsm"

    # Se convierten en data frame los archivos
    LISTA_POR_FINIQUITAR = pd.read_excel(RUTA_POR_FINIQUITAR, sheet_name="URL", header=None)
    BDD_MOD = pd.read_excel(RUTA_BUSCA_MODIFICATORIOS, sheet_name="M", header=0)

    # Se eliminan columnas sobrantes, columnas restantes dentro de la base de datos: ['Contrato', 'Título']
    BDD_MOD.drop(columns=['ID', 'Conjunto', 'Contratante', 'Concepto', 'contratista', \
                        'FirmaContrato', 'FechaInicio', 'FechaTermino', 'Fraccionamiento', \
                        'Importe', 'ContratoTipo', 'Descripcion', 'Satic', 'Creado por', \
                        'Creado', 'Etapa', 'superficie', 'conjunto_ID', \
                        'CantidadTrabajadoresContratista', 'Hipervinculo'], inplace=True)

    # Se revisa que contratos modificatorios existen dentro de la lista
    LISTA_EXISTENCIA = BDD_MOD['Contrato'].isin(LISTA_POR_FINIQUITAR[0]).rename('Existe')
    
    # Se concatena la lista que indíca si existe el contrato o no
    BDD_MOD = pd.concat([BDD_MOD, LISTA_EXISTENCIA], axis=1)

    # Filtramos la lista para que solo queden los existentes que son los que nos interesan
    BDD_MOD = BDD_MOD[BDD_MOD['Existe'] == True]

    # Tomamos cada finiquito de la lista finiquitar
    print(LISTA_POR_FINIQUITAR.columns)
    for FINIQUITO, CONTRATISTA in LISTA_POR_FINIQUITAR.values:
        
        ELEMENTOS = []

        # Buscamos el finquito dentro de la lista reducida de existentes
        MODIFICATORIOS = BDD_MOD[BDD_MOD['Contrato'] == FINIQUITO]['Título']
        ELEMENTOS.append(FINIQUITO + '_CF')

        for MODIFICATORIO in MODIFICATORIOS:
            ELEMENTOS.append(MODIFICATORIO)

        for ELEMENTO in ELEMENTOS:
            DRIVER.get("http://fprado@javer.com.mx:" + JU_PASSWORD + "@portal.javer.net/juridico/Contratos_Juridico/" + FINIQUITO + "/" + ELEMENTO + ".aspx")
            try:
                WebDriverWait(DRIVER, 2).until(EC.element_to_be_clickable((By.XPATH, "//img[@src='/_LAYOUTS/15/JAVERcontratos/img/LogoJaver_2020.png']")))
            except TimeoutException:
                try:
                    WebDriverWait(DRIVER, 2).until(EC.element_to_be_clickable((By.XPATH, "//img[@src='/_LAYOUTS/15/JAVERcontratos/img/logo_javer.png']")))
                except TimeoutException:
                    print(CONTRATISTA, ELEMENTO, " -NO EXISTÉ EN SISTEMA")
                    NO_EXISTE_SISTEMA.append("http://fprado@javer.com.mx:" + JU_PASSWORD + "@portal.javer.net/juridico/Contratos_Juridico/" + FINIQUITO + "/" + ELEMENTO + ".aspx")
                    break
            pyautogui.moveTo(683, 384)
            pyautogui.click()
            pyautogui.click()
            pyautogui.hotkey('ctrl', 'p')
            time.sleep(0.5)
            pyautogui.press('G')
            time.sleep(0.1)
            pyautogui.press('G')
            time.sleep(0.5)
            pyautogui.press('enter')
            time.sleep(0.5)
            pyautogui.write(CONTRATISTA + " - " + ELEMENTO, interval=0.0001)
            print(CONTRATISTA + "-" + ELEMENTO)
            time.sleep(0.25)
            pyautogui.press('enter')
            time.sleep(0.40)
            pyautogui.press('enter')
            time.sleep(0.40)
            pyautogui.press('enter')
            time.sleep(0.40)
            pyautogui.press('enter')
    print('Se acabó, estos no existen:')
    print(NO_EXISTE_SISTEMA)


#----FINIQUITO----------------------------------------------------------------------------------------------------------------------------------------------------


    # CONVERTIR EN DATA FRAME EL REPORTE
def dataframe_contratos():
    # inputs
    TODAY = str(datetime.date.today())
    FILE_NAME = "XXMCAN_XML_Report_Publisher_" + TODAY[8:] + TODAY[5:7] + TODAY[2:4] + ".xls"
    RUTA_DESCARGA = "C:\\Users\\fprado\\OneDrive - Servicios Administrativos Javer, S.A. DE C.V\\Descargas\\"
    RUTA_CARGA = "C:\\Users\\fprado\\REPORTES\\"
    RUTA_FILE = RUTA_DESCARGA + FILE_NAME
    #process
    DF_NC = pd.read_html(RUTA_FILE)[2]
    DF_NC = DF_NC[1:]
    DF_NC.columns = DF_NC.iloc[0]
    DF_NC = DF_NC[1:]
    print(DF_NC)
    DF_NCS = DF_NC.loc[DF_NC['Tipo documento'] == "SIROC"]
    DF_NCC = DF_NC.loc[DF_NC['Tipo documento'] != "SIROC"]
    with pd.ExcelWriter(RUTA_CARGA + 'CONCENTRADO_J.xlsx') as WRITER:
        DF_NCS.to_excel(WRITER, sheet_name='SIR')
        DF_NCC.to_excel(WRITER, sheet_name='CON')


def DESCARGA_REPCONT(driver):

    # DEFINICIONES CLAVE 
    TODAY = str(datetime.date.today())
    FILE_NAME = "XXMCAN_XML_Report_Publisher_" + TODAY[8:] + TODAY[5:7] + TODAY[2:4] + ".xls"
    RUTA_DESCARGA = "C:\\Users\\fprado\\OneDrive - Servicios Administrativos Javer, S.A. DE C.V\\Descargas\\"
    RUTA_CARGA = "C:\\Users\\fprado\\REPORTES\\"
    RUTA_FILE = RUTA_DESCARGA + FILE_NAME

    # INGRESAR A ORACLE Y DESCARGAR EL REPORTE
    acc_fir_oracle(driver)
    ir_en_fav(driver, "Sistema de reportes")
    ir_en_rg(driver, "Reporte de contratos legales [P]") 
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "ORIGEN_ID"))).send_keys("%", Keys.TAB, Keys.ENTER)
    frame_click_go(driver, "%Todos%")
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, "ButtonExecute"))).click()
    objetivo(driver, "XLS", "Reporte de contratos legales [P]")
    time.sleep(10)
    dataframe_contratos()
    time.sleep(4)
    driver.quit()


def BD_FINIQUITOS(driver):
    RECC(driver)
    RECC2 (driver)
    moveRECC()
    DB_FINIQUITO()


def BDDESC_REINT():
    D = DRIVER(2)
    acc_fir_oracle(D)
    lista = pd.read_csv("C:/Users/fprado/REPORTES/BDD/Requests.csv")
    listaa = lista["Requests"].tolist()
    print(listaa)
    descargarecc(D, listaa)


def DB_RDC_DOWN(driver):
    acc_fir_oracle(driver)
    ir_en_fav(driver, 'Sistema de reportes')
    ir_en_rg(driver, 'Reporte de contratos legales [P]')
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, "//table[@id='MainRN']/tbody[1]/tr[4]/td[1]/table[1]/tbody[1]/tr[1]/td[1]/div[1]/div[2]/table[1]/tbody[1]/tr[4]/td[1]/table[1]/tbody[1]/tr[1]/td[1]/div[1]/div[3]/span[1]/table[2]/tbody[1]/tr[7]/td[1]/a[1]/img[1]"))).click()
    frame_click_go(driver, "Actuales")
    objetivo(driver, "XLS", "Reporte de contratos legales [P]")


def moveRECC(A):
    filepath = "C:\\Users\\fprado\\OneDrive - Servicios Administrativos Javer, S.A. DE C.V\\Descargas"
    path = "C:\\Users\\fprado\\REPORTES\\REC\\"
    destination = [path + "RECC\\", path + "RDFC\\"]
    if A == "SI":
        for x in destination:
            for root, dirs, files in os.walk(x):
                for filename in files:
                    os.remove(x + filename)
    for root, dirs, files in os.walk(filepath):
        for filename in files:
            try:
                basename, extention = os.path.splitext(filename)
                if extention == '.xls' and basename[:26] == 'XXMCAN___Finiquito_de_Obra':
                    fullpath = root + '\\' + basename + extention
                    DOCNAM = destination[1] + basename + '.xls'
                    os.rename(fullpath, DOCNAM)
                elif extention == '.xls' and basename[:22] == 'XXMCAN__Reporte_Estado':
                    fullpath = root + '\\' + basename + extention
                    DOCNAM = destination[0] + basename + '.xls'
                    os.rename(fullpath, DOCNAM)
            except:
                print("no se pudo descargar")
                continue
    if not filepath =='':
        print("FINALIZADO")


def REQUERIMIENTO_ESTIMACIONES():
    filepath = "C:\\Users\\fprado\\REPORTES\\OneDrive_2_12-26-2022\\"
    destination = "C:\\Users\\fprado\\REPORTES\\REQUERIMIENTO_SAB\\"
    patron = ('.*E22.*', '.*E24.*', '.*E25.*')
    patron2 = ('.*U02.*', '.*I01.*')
    for root, dirs, files in os.walk(filepath):
        for filename in files:
            for pat in patron:
                if re.match(pat, filename):
                    for pat2 in patron2:
                        if re.match(pat2,filename):
                            break
                        else:
                            if re.match('.*CARATULA.*', root):
                                NOW = 'CARATULA\\'
                            else:
                                NOW = 'SOPORTE\\'
                            try:
                                fullpath = root + '\\' + filename
                                SEM = root[47:53]
                                newname = pat[2:5] + "\\" + "ESTIMACIONES\\" + NOW + SEM + " - " + filename
                                print(newname)
                                os.rename(fullpath, destination + newname)
                                time.sleep(.02)
                            except FileExistsError:
                                fullpath = root + '\\' + filename
                                SEM = root[47:53]
                                newname = pat[2:5] + "\\" + "ESTIMACIONES\\" + NOW + SEM + " - " + filename + "1"
                                print(newname)
                                os.rename(fullpath, destination + newname)
                            except FileExistsError:
                                fullpath = root + '\\' + filename
                                SEM = root[47:53]
                                newname = pat[2:5] + "\\" + "ESTIMACIONES\\" + NOW + SEM + " - " + filename + "2"
                                print(newname)
                                os.rename(fullpath, destination + newname)
                            except FileNotFoundError:
                                continue


def DyR_NUEVO_JUR(driver):
    acc_fir_oracle(driver)
    ir_en_fav(driver, "Sistema de reportes")
    ir_en_rg(driver, "Reporte de contratos legales [P]")
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH, PROYECTO))).click()
    frame_click_go(driver, proy)


def DB_FINIQUITO():
# RUTAS Y VARIABLES
    RUT = "C:\\Users\\fprado\\REPORTES\\"
    RUTAS = [RUT + "REC\\RDFC\\", RUT + "REC\\RECC\\"]
    RF_C, RF_L, RE_C, RE_L = [], [], [], []
# INGRESA CADA TABLA DE BASE DE DATOS EN UNA LISTA PARA SU POSTERIOR CONCATENACIÓN
    for RUTA in RUTAS:   
        LISC = glob.glob(RUTA + "/*.xls")
        for FILE in LISC:
            if RUTA[29:33] == "RDFC":
                TABLE = pd.read_html(FILE, header=0)
                RF_L.append(TABLE[1])
            elif RUTA[29:33] == "RECC":
                TABLE = pd.read_excel(FILE, header=1)
                RE_L.append(TABLE)
# CONCATENADOR DE LAS LISTAS EN DATA FRAMES PROVISIONALES             
    RDF = pd.concat(RF_L, ignore_index=True)
    REC = pd.concat(RE_L, ignore_index=True)
# DEPURACION DE DATA FRAMES DE DATOS INNECESARIOS
    RDF.dropna(subset=['Estado'], inplace=True)
    RDF.drop(columns=['Incurrido de MAT', 'Por incurrir MAT', 'Ahorros pendientes', 'Pendiente por contratar', 'Incurrido total', 'Por incurrir total', 'Estimado de cierre', 'Total Finiquito'], inplace=True)
    REC.dropna(subset=['Contrato'], inplace=True)
# COVERSIÓN A DATA FRAMES DE BASES DE DATOS EXTERNAS    
    CONC = pd.read_excel(RUT + 'BDD\\Contratos confirmados.xlsx', header=0)
    JUV = pd.read_excel(RUT + 'EJJ.xlsm', sheet_name='C', header=0, index_col=0)
    JUF = pd.read_excel(RUT + 'EJJ.xlsm', sheet_name='F', header=0, index_col=0)
    JUN = pd.read_excel(RUT + 'EJJ.xlsm', sheet_name='N', header=0, index_col=0)
# MEZCLA DE JURÍDICOS Y REVISIÓN DE EXISTENCIA DE JURÍDICOS
    JURC = pd.concat([JUV.filter(['Contrato', 'Conjunto', 'Contratista'], axis=1), JUN.filter(['Contrato', 'Conjunto', 'Contratista'], axis=1)])
# REVISIÓN DE EXISTENCIA DE CONTRATOS YA CONFIRMADOS CON ERROR Y CAMBIO POR LOS CONTRATOS CORRECTOS
    CONFI = RDF['Contrato'].isin(CONC['Contrato'])
    ESP =  RDF['Contrato'].map(dict(zip(CONC['Contrato'],CONC['Contrato Legal']))) 
    RDF['Contrato Legal'].mask(CONFI, ESP, inplace=True)
# ENCUENTRA LOS CONTRATOS QUE POSEÉN DESCRIPCIÓN
    ECRF = RDF['Contrato'].isin(REC['Contrato'])
    ECRF = ECRF.rename('Descripción')
    RDF = pd.concat([RDF, ECRF], axis=1)
    ESPF = RDF['Contrato'].map(dict(zip(REC['Contrato'],REC['Descripción.1'])))
    RDF['Descripción'].mask(ECRF, ESPF, inplace=True)
# LISTA DE EXISTENTES
    EXIST = RDF['Contrato Legal'].isin(JURC['Contrato'])
    EXIST = EXIST.rename('Existencia')
# EXTRACIÓN DE HIPERVÍNCULO - VIEJO SISTEMA
    HIPER = RDF['Contrato Legal'].isin(JUV['Contrato'])
    HIPER = HIPER.rename('Hipervínculo')
    RDF = pd.concat([RDF, HIPER], axis=1)
    HYPER = RDF['Contrato Legal'].map(dict(zip(JUV['Contrato'],JUV['Hipervinculo'])))
    RDF['Hipervínculo'].mask(HIPER, HYPER, inplace=True)
# LISTA DE FINIQUITADOS
    FINIQ = RDF['Contrato Legal'].isin(JUF['Contrato'])
    FINIQ = FINIQ.rename('Finiquitado')
# CONCATENA COLUMNA DE CODIGO DE ORACLE + CONTRATO LEGAL + EXISTENCIA EN UN DATAFRAME NUEVO BUSCA SU EXISTENCIA Y QUE NO SEA CARGO
    NEXIST = pd.concat([RDF['Contrato'], RDF['Contrato Legal'], EXIST], axis=1)
    NEXIST = NEXIST.loc[NEXIST['Existencia'] == False]
    NEXIST = NEXIST.loc[NEXIST['Contrato Legal'] != "CARGO"]
    # CONTCORR = NEXIST.filter(regex='*?QRO*?', axis=0)
    # print(CONTCORR)
# AGREGA AL DATA FRAME RDF SI YA ESTÁ FINIQUITADO EN LISTA Y LA COMPROBACIÓN DE SU EXISTENCIA
    RDF = pd.concat([RDF, FINIQ, EXIST], axis=1)
    RDF.replace(CONC, inplace=True)
    NEXIST.to_excel(RUT + "REC\\NEXIST.xlsx", index=False)
    RDF.to_excel(RUT + "REC\\RDF.xlsx", index=False)
    # REC.to_excel(RUT + "REC\\REC.xlsx", index=False)


def ARREGLO_DBF():
    X_EJ = []
    CONN = sqlite3.connect(':memory:')
    RUTA = "C:\\Users\\fprado\\REPORTES\\REC\\"
    X_EJ.append(pd.read_excel(RUTA[:25] + "EJJ.xlsm", sheet_name="C"))
    X_EJ.append(pd.read_excel(RUTA[:25] + "EJJ.xlsm", sheet_name="M"))
    X_EJ.append(pd.read_excel(RUTA[:25] + "EJJ.xlsm", sheet_name="F"))
    for EJ in X_EJ:
        FRACCIONAMIENTO = EJ.loc[EJ['Fraccionamiento']]
        print(FRACCIONAMIENTO)
        # EJ.drop(FRACCIONAMIENTO != 'PASEO SAN JUNIPERO' or FRACCIONAMIENTO != 'RANCHO EL SIETE' or FRACCIONAMIENTO != 'CUMBRE ALTA', inplace=True)
         # or EJ[EJ'Fraccionamiento'] != 'RANCHO EL SIETE' or EJ[EJ'Fraccionamiento'] != 'CUMBRE ALTA' or EJ[EJ'Fraccionamiento'] != 'MARQUES DEL RIO', inplace=True
        # print(EJ)
    X_RDF = pd.read_excel(RUTA + "RDF.xlsx")
    X_RDF.to_sql(name="DB_RDF", con=CONN, if_exists="append")


def DATABASEARR():
    RUTA = "C:\\Users\\fprado\\REPORTES\\REC\\"
    RUTAS = [RUTA + "RECC", RUTA + "RECC2"]
    REPORTE, REPORTE_LIST = [], [[], []]
    for CARPETA in RUTAS:
        V = RUTAS.index(CARPETA)
        LISTARC = glob.glob(RUTAS[V] + "/*.xls")
        for FILE in LISTARC:
            try:
                if V == 0:
                    TABLE = pd.read_excel(FILE, header=1)
                    REPORTE_LIST[V].append(TABLE)
                elif V == 1:
                    TABLE = pd.read_html(FILE, header=0)
                    REPORTE_LIST[V].append(TABLE[1])
            except:
                print("NO SE PUDO " + FILE)
        REPORTE.append(pd.concat(REPORTE_LIST[V], ignore_index=True))
    REC, RDF = REPORTE[0], REPORTE[1]
    REC.dropna(subset=['Fecha Inicial'], inplace=True)
    RDF.dropna(subset=['Estado'], inplace=True)
    RDF.drop(columns=['Incurrido de MAT', 'Por incurrir MAT', 'Ahorros pendientes', 'Pendiente por contratar', 'Incurrido total', 'Por incurrir total', 'Estimado de cierre', 'Total Finiquito'], inplace=True)
    REC.to_excel(RUTA + "REC.xlsx", index=False)
    RDF.to_excel(RUTA + "RDF.xlsx", index=False)

#---GLOBALES---
'''
HERE = os.getcwd() + "\\"


2 = 1
8 = 3


class ID():
    USER = ''
    PASSWORD = ''


ID = ID()
ID.USER = "FPRADO"
ID.PASSWORD = "PAMFYjaver07"


def create_driver(driver_type='firefox', headless=True, download_folder=None):
    """
    Esta función crea un driver, requisitos:
        driver_type= chrome, firefox o edge
        headless= True o False 
        download_folder = None o Ruta
    """
    if driver_type == 'chrome':
        options = ChromeOptions()
        if headless:
            options.add_argument('--headless')
        prefs = {'download.prompt_for_download': False,
                 'download.directory_upgrade': True,
                 'safebrowsing.enabled': False,
                 'plugins.always_open_pdf_externally': True}
        if download_folder:
            prefs['download.default_directory'] = download_folder
        options.add_experimental_option('prefs', prefs)
        return webdriver.Chrome(options=options)

    elif driver_type == 'firefox':
        options = FirefoxOptions()
        if headless:
            options.add_argument('-headless')
        options.set_preference("browser.download.folderList", 2)
        if download_folder:
            options.set_preference("browser.download.dir", download_folder)
        options.set_preference("browser.download.useDownloadDir", True)
        options.set_preference("browser.download.manager.showWhenStarting", False)
        options.set_preference("browser.helperApps.alwaysAsk.force", False)
        options.set_preference("browser.helperApps.neverAsk.saveToDisk", "application/pdf")
        options.set_preference("browser.download.manager.showAlertOnComplete", False)
        options.set_preference("browser.download.manager.useWindow", False)
        options.set_preference("pdfjs.disabled", True)
        options.set_preference("plugin.scan.plid.all", False)
        options.set_preference("dom.popup_maximum", 100)
        options.set_preference("app.update.enabled", False)
        return webdriver.Firefox(options=options)

    elif driver_type == 'edge':
        options = EdgeOptions()
        if headless:
            options.add_argument('headless')
        prefs = {'download.prompt_for_download': False,
                 'download.directory_upgrade': True,
                 'safebrowsing.enabled': False,
                 'plugins.always_open_pdf_externally': True}
        if download_folder:
            prefs['download.default_directory'] = download_folder
        options.set_capability("prefs", prefs)
        return webdriver.Edge(options=options)

''
def DRI(choose):
    DEFINE EL driver A USAR
        LOS driverS SON:
        1 = HIDEN FIREFOX
        2  = FIREFOX
        3 = HIDEN CHROME
        4  = CHROME
    path = 'C://Users/fporado/Onedriverve - Servicios Administrativos Javer, S.A. DE C.V/Escritorio/PROTECCIÓN/'
    extens = 'application/vnd.hzn-3d-crossword,video/3gpp,video/3gpp2,application/vnd.mseq,application/vnd.3m.post-it-notes,application/vnd.3gpp.pic-bw-large,application/vnd.3gpp.pic-bw-small,application/vnd.3gpp.pic-bw-var,application/vnd.3gp2.tcap,application/x-7z-compressed,application/x-abiword,application/x-ace-compressed,application/vnd.americandynamics.acc,application/vnd.acucobol,application/vnd.acucorp,audio/adpcm,application/x-authorware-bin,application/x-athorware-map,application/x-authorware-seg,application/vnd.adobe.air-application-installer-package+zip,application/x-shockwave-flash,application/vnd.adobe.fxp,application/pdf,application/vnd.cups-ppd,application/x-director,applicaion/vnd.adobe.xdp+xml,application/vnd.adobe.xfdf,audio/x-aac,application/vnd.ahead.space,application/vnd.airzip.filesecure.azf,application/vnd.airzip.filesecure.azs,application/vnd.amazon.ebook,application/vnd.amiga.ami,applicatin/andrew-inset,application/vnd.android.package-archive,application/vnd.anser-web-certificate-issue-initiation,application/vnd.anser-web-funds-transfer-initiation,application/vnd.antix.game-component,application/vnd.apple.installe+xml,application/applixware,application/vnd.hhe.lesson-player,application/vnd.aristanetworks.swi,text/x-asm,application/atomcat+xml,application/atomsvc+xml,application/atom+xml,application/pkix-attr-cert,audio/x-aiff,video/x-msvieo,application/vnd.audiograph,image/vnd.dxf,model/vnd.dwf,text/plain-bas,application/x-bcpio,application/octet-stream,image/bmp,application/x-bittorrent,application/vnd.rim.cod,application/vnd.blueice.multipass,application/vnd.bm,application/x-sh,image/prs.btif,application/vnd.businessobjects,application/x-bzip,application/x-bzip2,application/x-csh,text/x-c,application/vnd.chemdraw+xml,text/css,chemical/x-cdx,chemical/x-cml,chemical/x-csml,application/vn.contact.cmsg,application/vnd.claymore,application/vnd.clonk.c4group,image/vnd.dvb.subtitle,application/cdmi-capability,application/cdmi-container,application/cdmi-domain,application/cdmi-object,application/cdmi-queue,applicationvnd.cluetrust.cartomobile-config,application/vnd.cluetrust.cartomobile-config-pkg,image/x-cmu-raster,model/vnd.collada+xml,text/csv,application/mac-compactpro,application/vnd.wap.wmlc,image/cgm,x-conference/x-cooltalk,image/x-cmx,application/vnd.xara,application/vnd.cosmocaller,application/x-cpio,application/vnd.crick.clicker,application/vnd.crick.clicker.keyboard,application/vnd.crick.clicker.palette,application/vnd.crick.clicker.template,application/vn.crick.clicker.wordbank,application/vnd.criticaltools.wbs+xml,application/vnd.rig.cryptonote,chemical/x-cif,chemical/x-cmdf,application/cu-seeme,application/prs.cww,text/vnd.curl,text/vnd.curl.dcurl,text/vnd.curl.mcurl,text/vnd.crl.scurl,application/vnd.curl.car,application/vnd.curl.pcurl,application/vnd.yellowriver-custom-menu,application/dssc+der,application/dssc+xml,application/x-debian-package,audio/vnd.dece.audio,image/vnd.dece.graphic,video/vnd.dec.hd,video/vnd.dece.mobile,video/vnd.uvvu.mp4,video/vnd.dece.pd,video/vnd.dece.sd,video/vnd.dece.video,application/x-dvi,application/vnd.fdsn.seed,application/x-dtbook+xml,application/x-dtbresource+xml,application/vnd.dvb.ait,applcation/vnd.dvb.service,audio/vnd.digital-winds,image/vnd.djvu,application/xml-dtd,application/vnd.dolby.mlp,application/x-doom,application/vnd.dpgraph,audio/vnd.dra,application/vnd.dreamfactory,audio/vnd.dts,audio/vnd.dts.hd,imag/vnd.dwg,application/vnd.dynageo,application/ecmascript,application/vnd.ecowin.chart,image/vnd.fujixerox.edmics-mmr,image/vnd.fujixerox.edmics-rlc,application/exi,application/vnd.proteus.magazine,application/epub+zip,message/rfc82,application/vnd.enliven,application/vnd.is-xpr,image/vnd.xiff,application/vnd.xfdl,application/emma+xml,application/vnd.ezpix-album,application/vnd.ezpix-package,image/vnd.fst,video/vnd.fvt,image/vnd.fastbidsheet,application/vn.denovo.fcselayout-link,video/x-f4v,video/x-flv,image/vnd.fpox,image/vnd.net-fpox,text/vnd.fmi.flexstor,video/x-fli,application/vnd.fluxTIME.clip,application/vnd.fdf,text/x-fortran,application/vnd.mif,application/vnd.framemaker,imae/x-freehand,application/vnd.fsc.weblaunch,application/vnd.frogans.fnc,application/vnd.frogans.ltf,application/vnd.fujixerox.ddd,application/vnd.fujixerox.docuworks,application/vnd.fujixerox.docuworks.binder,application/vnd.fujitu.oasys,application/vnd.fujitsu.oasys2,application/vnd.fujitsu.oasys3,application/vnd.fujitsu.oasysgp,application/vnd.fujitsu.oasysprs,application/x-futuresplash,application/vnd.fuzzysheet,image/g3fax,application/vnd.gmx,model/vn.gtw,application/vnd.genomatix.tuxedo,application/vnd.geogebra.file,application/vnd.geogebra.tool,model/vnd.gdl,application/vnd.geometry-explorer,application/vnd.geonext,application/vnd.geoplan,application/vnd.geospace,applicatio/x-font-ghostscript,application/x-font-bdf,application/x-gtar,application/x-texinfo,application/x-gnumeric,application/vnd.google-earth.kml+xml,application/vnd.google-earth.kmz,application/vnd.grafeq,image/gif,text/vnd.graphviz,aplication/vnd.groove-account,application/vnd.groove-help,application/vnd.groove-identity-message,application/vnd.groove-injector,application/vnd.groove-tool-message,application/vnd.groove-tool-template,application/vnd.groove-vcar,video/h261,video/h263,video/h264,application/vnd.hp-hpid,application/vnd.hp-hps,application/x-hdf,audio/vnd.rip,application/vnd.hbci,application/vnd.hp-jlyt,application/vnd.hp-pcl,application/vnd.hp-hpgl,application/vnd.yamaha.h-script,application/vnd.yamaha.hv-dic,application/vnd.yamaha.hv-voice,application/vnd.hydrostatix.sof-data,application/hyperstudio,application/vnd.hal+xml,text/html,application/vnd.ibm.rights-management,application/vnd.ibm.securecontainer,text/calendar,application/vnd.iccprofile,image/x-icon,application/vnd.igloader,image/ief,application/vnd.immervision-ivp,application/vnd.immervision-ivu,application/reginfo+xml,text/vnd.in3d.3dml,text/vnd.in3d.spot,mode/iges,application/vnd.intergeo,application/vnd.cinderella,application/vnd.intercon.formnet,application/vnd.isac.fcs,application/ipfix,application/pkix-cert,application/pkixcmp,application/pkix-crl,application/pkix-pkipath,applicaion/vnd.insors.igm,application/vnd.ipunplugged.rcprofile,application/vnd.irepository.package+xml,text/vnd.sun.j2me.app-descriptor,application/java-archive,application/java-vm,application/x-java-jnlp-file,application/java-serializd-object,text/x-java-source,java,application/javascript,application/json,application/vnd.joost.joda-archive,video/jpm,image/jpeg,video/jpeg,application/vnd.kahootz,application/vnd.chipnuts.karaoke-mmd,application/vnd.kde.karbon,aplication/vnd.kde.kchart,application/vnd.kde.kformula,application/vnd.kde.kivio,application/vnd.kde.kontour,application/vnd.kde.kpresenter,application/vnd.kde.kspread,application/vnd.kde.kword,application/vnd.kenameaapp,applicatin/vnd.kidspiration,application/vnd.kinar,application/vnd.kodak-descriptor,application/vnd.las.las+xml,application/x-latex,application/vnd.llamagraphics.life-balance.desktop,application/vnd.llamagraphics.life-balance.exchange+xml,application/vnd.jam,application/vnd.lotus-1-2-3,application/vnd.lotus-approach,application/vnd.lotus-freelance,application/vnd.lotus-notes,application/vnd.lotus-organizer,application/vnd.lotus-screencam,application/vnd.lotus-wordro,audio/vnd.lucent.voice,audio/x-mpegurl,video/x-m4v,application/mac-binhex40,application/vnd.macports.portpkg,application/vnd.osgeo.mapguide.package,application/marc,application/marcxml+xml,application/mxf,application/vnd.wolfrm.player,application/mathematica,application/mathml+xml,application/mbox,application/vnd.medcalcdata,application/mediaservercontrol+xml,application/vnd.mediastation.cdkey,application/vnd.mfer,application/vnd.mfmp,model/mesh,appliation/mads+xml,application/mets+xml,application/mods+xml,application/metalink4+xml,application/vnd.ms-powerpoint.template.macroenabled.12,application/vnd.ms-word.document.macroenabled.12,application/vnd.ms-word.template.macroenabed.12,application/vnd.mcd,application/vnd.micrografx.flo,application/vnd.micrografx.igx,application/vnd.eszigno3+xml,application/x-msaccess,video/x-ms-asf,application/x-msdownload,application/vnd.ms-artgalry,application/vnd.ms-ca-compressed,application/vnd.ms-ims,application/x-ms-application,application/x-msclip,image/vnd.ms-modi,application/vnd.ms-fontobject,application/vnd.ms-excel,application/vnd.ms-excel.addin.macroenabled.12,application/vnd.ms-excelsheet.binary.macroenabled.12,application/vnd.ms-excel.template.macroenabled.12,application/vnd.ms-excel.sheet.macroenabled.12,application/vnd.ms-htmlhelp,application/x-mscardfile,application/vnd.ms-lrm,application/x-msmediaview,aplication/x-msmoney,application/vnd.openxmlformats-officedocument.presentationml.presentation,application/vnd.openxmlformats-officedocument.presentationml.slide,application/vnd.openxmlformats-officedocument.presentationml.slideshw,application/vnd.openxmlformats-officedocument.presentationml.template,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/vnd.openxmlformats-officedocument.spreadsheetml.template,application/vnd.openxmformats-officedocument.wordprocessingml.document,application/vnd.openxmlformats-officedocument.wordprocessingml.template,application/x-msbinder,application/vnd.ms-officetheme,application/onenote,audio/vnd.ms-playready.media.pya,vdeo/vnd.ms-playready.media.pyv,application/vnd.ms-powerpoint,application/vnd.ms-powerpoint.addin.macroenabled.12,application/vnd.ms-powerpoint.slide.macroenabled.12,application/vnd.ms-powerpoint.presentation.macroenabled.12,appliation/vnd.ms-powerpoint.slideshow.macroenabled.12,application/vnd.ms-project,application/x-mspublisher,application/x-msschedule,application/x-silverlight-app,application/vnd.ms-pki.stl,application/vnd.ms-pki.seccat,application/vn.visio,video/x-ms-wm,audio/x-ms-wma,audio/x-ms-wax,video/x-ms-wmx,application/x-ms-wmd,application/vnd.ms-wpl,application/x-ms-wmz,video/x-ms-wmv,video/x-ms-wvx,application/x-msmetafile,application/x-msterminal,application/msword,application/x-mswrite,application/vnd.ms-works,application/x-ms-xbap,application/vnd.ms-xpsdocument,audio/midi,application/vnd.ibm.minipay,application/vnd.ibm.modcap,application/vnd.jcp.javame.midlet-rms,application/vnd.tmobile-ivetv,application/x-mobipocket-ebook,application/vnd.mobius.mbk,application/vnd.mobius.dis,application/vnd.mobius.plc,application/vnd.mobius.mqy,application/vnd.mobius.msl,application/vnd.mobius.txf,application/vnd.mobius.daf,tex/vnd.fly,application/vnd.mophun.certificate,application/vnd.mophun.application,video/mj2,audio/mpeg,video/vnd.mpegurl,video/mpeg,application/mp21,audio/mp4,video/mp4,application/mp4,application/vnd.apple.mpegurl,application/vnd.msician,application/vnd.muvee.style,application/xv+xml,application/vnd.nokia.n-gage.data,application/vnd.nokia.n-gage.symbian.install,application/x-dtbncx+xml,application/x-netcdf,application/vnd.neurolanguage.nlu,application/vnd.na,application/vnd.noblenet-directory,application/vnd.noblenet-sealer,application/vnd.noblenet-web,application/vnd.nokia.radio-preset,application/vnd.nokia.radio-presets,text/n3,application/vnd.novadigm.edm,application/vnd.novadim.edx,application/vnd.novadigm.ext,application/vnd.flographit,audio/vnd.nuera.ecelp4800,audio/vnd.nuera.ecelp7470,audio/vnd.nuera.ecelp9600,application/oda,application/ogg,audio/ogg,video/ogg,application/vnd.oma.dd2+xml,applicatin/vnd.oasis.opendocument.text-web,application/oebps-package+xml,application/vnd.intu.qbo,application/vnd.openofficeorg.extension,application/vnd.yamaha.openscoreformat,audio/webm,video/webm,application/vnd.oasis.opendocument.char,application/vnd.oasis.opendocument.chart-template,application/vnd.oasis.opendocument.database,application/vnd.oasis.opendocument.formula,application/vnd.oasis.opendocument.formula-template,application/vnd.oasis.opendocument.grapics,application/vnd.oasis.opendocument.graphics-template,application/vnd.oasis.opendocument.image,application/vnd.oasis.opendocument.image-template,application/vnd.oasis.opendocument.presentation,application/vnd.oasis.opendocumen.presentation-template,application/vnd.oasis.opendocument.spreadsheet,application/vnd.oasis.opendocument.spreadsheet-template,application/vnd.oasis.opendocument.text,application/vnd.oasis.opendocument.text-master,application/vnd.asis.opendocument.text-template,image/ktx,application/vnd.sun.xml.calc,application/vnd.sun.xml.calc.template,application/vnd.sun.xml.draw,application/vnd.sun.xml.draw.template,application/vnd.sun.xml.impress,application/vnd.sun.xl.impress.template,application/vnd.sun.xml.math,application/vnd.sun.xml.writer,application/vnd.sun.xml.writer.global,application/vnd.sun.xml.writer.template,application/x-font-otf,application/vnd.yamaha.openscoreformat.osfpovg+xml,application/vnd.osgi.dp,application/vnd.palm,text/x-pascal,application/vnd.pawaafile,application/vnd.hp-pclxl,application/vnd.picsel,image/x-pcx,image/vnd.adobe.photoshop,application/pics-rules,image/x-pict,application/x-chat,aplication/pkcs10,application/x-pkcs12,application/pkcs7-mime,application/pkcs7-signature,application/x-pkcs7-certreqresp,application/x-pkcs7-certificates,application/pkcs8,application/vnd.pocketlearn,image/x-portable-anymap,image/-portable-bitmap,application/x-font-pcf,application/font-tdpfr,application/x-chess-pgn,image/x-portable-graymap,image/png,image/x-portable-pixmap,application/pskc+xml,application/vnd.ctc-posml,application/postscript,application/xfont-type1,application/vnd.powerbuilder6,application/pgp-encrypted,application/pgp-signature,application/vnd.previewsystems.box,application/vnd.pvi.ptid1,application/pls+xml,application/vnd.pg.format,application/vnd.pg.osasli,tex/prs.lines.tag,application/x-font-linux-psf,application/vnd.publishare-delta-tree,application/vnd.pmi.widget,application/vnd.quark.quarkxpress,application/vnd.epson.esf,application/vnd.epson.msf,application/vnd.epson.ssf,applicaton/vnd.epson.quickanime,application/vnd.intu.qfx,video/quickTIME,application/x-rar-compressed,audio/x-pn-realaudio,audio/x-pn-realaudio-plugin,application/rsd+xml,application/vnd.rn-realmedia,application/vnd.realvnc.bed,applicatin/vnd.recordare.musicxml,application/vnd.recordare.musicxml+xml,application/relax-ng-compact-syntax,application/vnd.data-vision.rdz,application/rdf+xml,application/vnd.cloanto.rp9,application/vnd.jisp,application/rtf,text/richtex,application/vnd.route66.link66+xml,application/rss+xml,application/shf+xml,application/vnd.sailingtracker.track,image/svg+xml,application/vnd.sus-calendar,application/sru+xml,application/set-payment-initiation,application/set-reistration-initiation,application/vnd.sema,application/vnd.semd,application/vnd.semf,application/vnd.seemail,application/x-font-snf,application/scvp-vp-request,application/scvp-vp-response,application/scvp-cv-request,application/svp-cv-response,application/sdp,text/x-setext,video/x-sgi-movie,application/vnd.shana.informed.formdata,application/vnd.shana.informed.formtemplate,application/vnd.shana.informed.interchange,application/vnd.shana.informed.package,application/thraud+xml,application/x-shar,image/x-rgb,application/vnd.epson.salt,application/vnd.accpac.simply.aso,application/vnd.accpac.simply.imp,application/vnd.simtech-mindmapper,application/vnd.commonspace,application/vnd.ymaha.smaf-audio,application/vnd.smaf,application/vnd.yamaha.smaf-phrase,application/vnd.smart.teacher,application/vnd.svd,application/sparql-query,application/sparql-results+xml,application/srgs,application/srgs+xml,application/sml+xml,application/vnd.koan,text/sgml,application/vnd.stardivision.calc,application/vnd.stardivision.draw,application/vnd.stardivision.impress,application/vnd.stardivision.math,application/vnd.stardivision.writer,application/vnd.tardivision.writer-global,application/vnd.stepmania.stepchart,application/x-stuffit,application/x-stuffitx,application/vnd.solent.sdkm+xml,application/vnd.olpc-sugar,audio/basic,application/vnd.wqd,application/vnd.symbian.install,application/smil+xml,application/vnd.syncml+xml,application/vnd.syncml.dm+wbxml,application/vnd.syncml.dm+xml,application/x-sv4cpio,application/x-sv4crc,application/sbml+xml,text/tab-separated-values,image/tiff,application/vnd.to.intent-module-archive,application/x-tar,application/x-tcl,application/x-tex,application/x-tex-tfm,application/tei+xml,text/plain,application/vnd.spotfire.dxp,application/vnd.spotfire.sfs,application/TIMEstamped-data,applicationvnd.trid.tpt,application/vnd.triscape.mxs,text/troff,application/vnd.trueapp,application/x-font-ttf,text/turtle,application/vnd.umajin,application/vnd.uoml+xml,application/vnd.unity,application/vnd.ufdl,text/uri-list,application/nd.uiq.theme,application/x-ustar,text/x-uuencode,text/x-vcalendar,text/x-vcard,application/x-cdlink,application/vnd.vsf,model/vrml,application/vnd.vcx,model/vnd.mts,model/vnd.vtu,application/vnd.visionary,video/vnd.vivo,applicatin/ccxml+xml,,application/voicexml+xml,application/x-wais-source,application/vnd.wap.wbxml,image/vnd.wap.wbmp,audio/x-wav,application/davmount+xml,application/x-font-woff,application/wspolicy+xml,image/webp,application/vnd.webturb,application/widget,application/winhlp,text/vnd.wap.wml,text/vnd.wap.wmlscript,application/vnd.wap.wmlscriptc,application/vnd.wordperfect,application/vnd.wt.stf,application/wsdl+xml,image/x-xbitmap,image/x-xpixmap,image/x-xwindowump,application/x-x509-ca-cert,application/x-xfig,application/xhtml+xml,application/xml,application/xcap-diff+xml,application/xenc+xml,application/patch-ops-error+xml,application/resource-lists+xml,application/rls-services+xml,aplication/resource-lists-diff+xml,application/xslt+xml,application/xop+xml,application/x-xpinstall,application/xspf+xml,application/vnd.mozilla.xul+xml,chemical/x-xyz,text/yaml,application/yang,application/yin+xml,application/vnd.ul,application/zip,application/vnd.handheld-entertainment+xml,application/vnd.zzazz.deck+xml'
    #  FIREFOX PROFILE ESCONDIDO, DESCARGA AUTOMÁTICA
    fph = FO()
    fph.add_argument('-headless')
    fph.set_preference("browser.download.folderList", 2)
    fph.set_preference("browser.download.dir", path)
    fph.set_preference("browser.download.useDownloadDir", True)
    fph.set_preference("browser.download.manager.showWhenStarting", False)
    fph.set_preference("browser.helperApps.alwaysAsk.force", False)
    fph.set_preference("browser.helperApps.neverAsk.saveToDisk", extens)
    fph.set_preference("browser.download.manager.showAlertOnComplete", False)
    fph.set_preference("browser.download.manager.useWindow", False)
    fph.set_preference("pdfjs.disabled", True)
    fph.set_preference("plugin.scan.plid.all", False)
    fph.set_preference("dom.popup_maximum", 100)
    fph.set_preference("app.update.enabled", False)
    #   FIREFOX PROFILE VENTANA ABIERTA, DESCARGA AUTOMÁTICA
    fpo = FO()
    fpo.set_preference("app.update.enabled", False)
    fpo.set_preference("browser.download.folderList", 2)
    fpo.set_preference("browser.download.dir", path)
    fpo.set_preference("browser.download.useDownloadDir", True)
    fpo.set_preference("browser.download.manager.showWhenStarting", False)
    fpo.set_preference("browser.helperApps.alwaysAsk.force", False)
    fpo.set_preference("browser.helperApps.neverAsk.saveToDisk", extens)
    fpo.set_preference("browser.download.manager.showAlertOnComplete", False)
    fpo.set_preference("browser.download.manager.useWindow", False)
    fpo.set_preference("pdfjs.disabled", True)
    fpo.set_preference("plugin.scan.plid.all", False)
    fpo.set_preference("dom.popup_maximum", 100)
    #   CHROME PROFILE VENTANA ESCONDIDA DESCARGA AUTOMÁTICA
    cph = webdriver.ChromeOptions()
    profile = { "download.prompt_for_download": False}
    cph.add_experimental_option("prefs", profile)
    cph.add_argument("--headless")
    #   CHROME PROFILE VENTNA ABIERTA DESCARGA AUTOMÁTICA
    cpo = webdriver.ChromeOptions()
    profile = { "download.prompt_for_download": False}
    cpo.add_experimental_option("prefs", profile)
    #   UBICACIÓN DE driverS
    GECKODV = FS(os.getcwd() + "\\geckodriver.exe")
    CHROMDV = CS(os.getcwd() + "\\chromedriver.exe")
    #   driver CHOOSE
    if choose == 1:
        driver = webdriver.Firefox(options=fph, service=GECKODV)
    elif choose == 2:
        driver = webdriver.Firefox(options=fpo, service=GECKODV)
    elif choose == 3:
        driver = webdriver.Chrome(service=CHROMDV)
    elif choose == 4:
        driver = webdriver.Chrome(service=CHROMDV, options=cpo)
    return driver


def convertir_conjunto(CONJUNTO):
    DADO UN CONJUNTO DEVUELVE EL: FRACCIONAMIENTO, FRENTE, ETAPA
    E00 = ["E01", "E02", "E03", "E04", "E05", "E06", "E07", "E08", "E09", "E10", "E11", "E12", "E13", "E14", "E15", "E16", "E17", "E18", "E19", "E20", "E21", "E22", "E23", "E24", "E25", "E26", "E27"]
    UJN = ["VSU", "   ", "RLU", "BSU", "S2U", "BDU", "P2U", "   ", "   ", "USV", "ELU", "UED", "UB4", "   ", "URC", "UNL", "JUU", "USI", "   ", "UR7", "UMA", "UFB", "UPR", "UMO", "CJM", "   ", "   "]
    CJQ = ["CST", "   ", "CRL", "CB2", "CS2", "CB3", "CHM", "   ", "   ", "CS3", "CEL", "CVP", "CB4", "   ", "CRÑ", "CMN", "CBJ", "CSI", "   ", "CR7", "CMA", "CFB", "CPR", "CMS", "CJM", "   ", "PME"]
    #SCJ = 
    PRIMER = CONJUNTO[0:3]
    ETAPA = CONJUNTO[7:10]
    FRENTE = CONJUNTO[4:6]
    NUMERO = int(PRIMER[1:3]) - 1
    if ETAPA == "I01" or ETAPA == "U02":
        if int(FRENTE) > 79:
            FRAC = CJQ[NUMERO]
        else:
            FRAC = UJN[NUMERO]
    else: 
        FRAC = CJQ[NUMERO]
    return (FRAC, FRENTE, ETAPA)

#----CREAR ORGANIZACIONES EN EXCEL---

def ORGANIZACIONES_TO_XLSX(LISTA):

    DF_ORGANIZACIONES = pd.DataFrame(LISTA, columns=["Organizaciones"])
    DF_ORGANIZACIONES.to_excel('BDD\\BDD_ORGANIZACIONES.xlsx', index=False)


#----OBJETOS QUE REDUCEN ACCIONES----
 
def swdw(driver, tiempo, LOOKBY, LOOKFOR):

    Da click al objeto seleccionado: driver, tiempo, que buscas
    
    Args: 
        driver: el driver a usar
        TIME: el tiempo a esperar
        LOOKBY: 0: By.XPATH, 1: By.ID, 2: By.CSS_SELECTOR, 3: By.NAME, 4: By.PARTIAL_LINK_TEXT
        LOOKFOR: EL CÓDIGO A BUSCAR

    # Forma de busqueda
    buscar_por = {0: By.XPATH, 1: By.ID, 2: By.CSS_SELECTOR, 3: By.NAME, 4: By.PARTIAL_LINK_TEXT}
    fragmentos = 6
    tiempito = (tiempo + 1) / (fragmentos)
    for C in range(fragmentos):
        if C == (fragmentos - 1):
            while True:
                try:
                    SIMPLE = WebDriverWait(driver, tiempo).until(EC.element_to_be_clickable((buscar_por[LOOKBY], LOOKFOR)))
                    break
                except (NoSuchWindowException):
                    continue
            break
        try:
            SIMPLE = WebDriverWait(driver, tiempito).until(EC.element_to_be_clickable((buscar_por[LOOKBY], LOOKFOR)))
            break
        except TimeoutException:
            time.sleep(0.2)
            continue
        except (StaleElementReferenceException, NoSuchWindowException):
            time.sleep(1)
            continue
    return SIMPLE


def stf(driver, LOOKBY, LOOKFOR, SEND):
DENTRO DEL MODULO DE CONSTRUCCIÓN INTRODUCE UN DATO DENTRO DEL FRAME
    while True:
        # Hacemos click en el elemento que querémos abrir en frame
        try:
            driver.switch_to.window(driver.window_handles[0])
            swdw(driver, 3, LOOKBY, LOOKFOR).click()
            # Revisamos que se abra otra ventana
            ventanas = driver.window_handles
            cantidad_ventanas = len(ventanas)
            time.sleep(0.8)
            if cantidad_ventanas == 2:
                break
            else:
                time.sleep(0.5)
                continue
        except (TimeoutException, StaleElementReferenceException):
            continue
        except ElementClickInterceptedException:
            time.sleep(1)
            ventanas = driver.window_handles
            cantidad_ventanas = len(ventanas)
            if cantidad_ventanas == 2:
                break
            else:
                time.sleep(0.5)
                continue
    while True:
        for counter in range(10):
            try:
                driver.switch_to.window(driver.window_handles[1])
                driver.switch_to.frame(0)
                break
            except (NoSuchFrameException, IndexError, NoSuchWindowException):
                time.sleep(0.8)
                continue
        try:
            swdw(driver, 5, 0, "//input[@title='Término de Búsqueda']").clear()
            swdw(driver, 2, 0, "//input[@title='Término de Búsqueda']").send_keys("%" + SEND + "%" + Keys.TAB)
            swdw(driver, 2, 0, "//button[text()='Ir']").click()
            swdw(driver, 1, 0, "//table[@class='x1o']/tbody[1]/tr[2]/td[2]/a[1]/img[1]").click()
            time.sleep(0.8)
        except TimeoutException:
            continue
        except (NoSuchWindowException, IndexError):
            break
        ventanas = driver.window_handles
        cantidad_ventanas = len(ventanas)
        if cantidad_ventanas == 1:
            driver.switch_to.window(driver.window_handles[0])
            driver.switch_to.default_content()
            swdw(driver, 4, 0, '//body').click()
            break


def beautiful_table(driver, element="class", name="x1o"):

    # Almacenadores
    headers = []
    rows = []

    # Obtener el contenido HTML de la página web
    html = driver.page_source
    # Crear un objeto BeautifulSoup a partir del HTML
    soup = BeautifulSoup(html, 'html.parser')    
    # Encontrar la tabla en el HTML
    table = soup.find('table', {element: name})
   
    # Extraer los encabezados
    header_row = table.find('tr')
    if header_row:
        list_head = header_row.find_all('th')
        for th in list_head:
            text = th.text.strip()
            idd = th.get('id')
            indexx = list_head.index(th) 
            if text:
                headers.append(text)
            elif idd:
                headers.append(idd)
            else:
                headers.append(indexx)
   
    for row in table.find_all('tr')[1:]:
        cells = []
        for td in row.find_all('td'):
            span = td.find('span')  # Buscar solo el primer span en la celda
            a = td.find_all('a')
            if len(a) == 1:
                a = a[0]
            elif len(a) > 1:
                a = a[1]

            # 
            if span:
                text = span.text.strip()
                idd = span.get('id')
            elif a:
                text = a.text
                idd = a.get('id')
            else:
                text = td.text
                idd = td.get('id')

            # Se asigna el valor
            if text != "":
                cells.append(text)
            elif idd != "":
                cells.append(idd)
            else:
                cells.append(td)
        rows.append(cells)

    # Pasamos a DataFrame
    df = pd.DataFrame(rows, columns=headers)
    
    return df


#------------------------------------------------OBJETOS BASE PARA RUTAS PREDEFINIDA---------------

def acceder_oracle(driver):
    DADO UN USUARIO Y CONTRASEÑA INGRESA EN ORACLE
    user = ID.USER
    password = ID.PASSWORD
    HOME = "//table[@class='x6w']/tbody[1]/tr[1]/td[3]/a[1]"
    driver.get("http://siapp3.javer.com.mx:8010/OA_HTML/AppsLogin")
    if swdw(driver, 10, 0, "//table[@id='langOptionsTable']/tbody[1]/tr[2]/td[2]/span[1]").text == "Select a Language:":
        swdw(driver, 1, 0, "//img[@title='Latin American Spanish']").click()
    swdw(driver, 10, 3, "usernameField").clear()
    swdw(driver, 1, 3, "usernameField").send_keys(user + Keys.TAB)
    swdw(driver, 1, 3, "passwordField").send_keys(password + Keys.ENTER)
    swdw(driver, 10, 0, HOME).click()


def new_mainmenu(driver, ruta):

        da click a los elementos requeridos.
    # Revizamos estar en mainmenu buscando un elemento en caso de no estarlo damos click a ir al main menu
    try:
        swdw(driver, 1, 0, "//table[@id='respList']/tbody[1]/tr[1]/td[1]/ul[1]/li[1]/a[1]")
    except TimeoutException:
        swdw(driver, 2, 0, "//table[@class='x6w']/tbody[1]/tr[1]/td[3]/a[1]").click()

    # Tratamos de dar click al último elemento, si no, vamos uno por uno
    try:
        xpath = "//a[contains(text(), '" + ruta[-1] + "')]"
        swdw(driver, 2, 0, xpath).click()
    except TimeoutException:
        for element in ruta:
            xpath = "//a[contains(text(), '" + element + "')]"
            swdw(driver, 2, 0, xpath).click()
    except UnexpectedAlertPresentException:
        # Si se lanza la excepción UnexpectedAlertPresentException, significa que hay una alerta presente.
        try:
            # Manejar la alerta haciendo clic en el botón "Aceptar" o "Cancelar" según corresponda
            alert = driver.switch_to.alert
            alert.accept()
        except NoAlertPresentException:
            time.sleep(0.1)


#------------------------------------------------HERRAMIENTAS DE BUSQUEDA--------------------------

def go_front(driver):

    while True:

        # Definiciones
        buscar_frentes = ["JAV_MC_CAO_QRO", "Frentes", "Buscar frentes"]

        try:
            # Verificador de estar en frentes
            swdw(driver, 10, 0, "//h2[text()='Frentes de Construcción']")
            break

        except TimeoutException:

            # Intentamos click en buscar frentes
            try:
                new_mainmenu(driver, buscar_frentes)

            # De no poder reingresamos
            except TimeoutException:
                acceder_oracle(driver)
                new_mainmenu(driver, buscar_frentes)


def go_sistema_reportes(driver):

    while True:

        # Definiciones
        buscar = ["JAV_MC_CAO_QRO", "Reportes Generales", "Sistema de reportes"]

        try:
            # Verificador de sistema de reportes
            swdw(driver, 2, 0, "(//span[text()='Hacer clic en el icono para ejecutar el reporte'])[1]")
            break

        except TimeoutException:

            # Intentamos click en buscar frentes
            try:
                new_mainmenu(driver, buscar)

            # De no poder reingresamos
            except TimeoutException:
                acceder_oracle(driver)
                new_mainmenu(driver, buscar)


def go_org(driver, ORG):

    # Definiciones
    xpath_org = "(//span[@id='OrganizationLOV__xc_0']//img)[2]"

    while True:
        # Intentamos dos veces
        try:
            # Enviamos la organización al buscador de frentes y damos click
            stf(driver, 0, xpath_org, ORG)
            swdw(driver, 8, 1, "OrganizationLOV").click()
            swdw(driver, 2, 0, "//button[@id='Search']").click()

            # Ampliamos busqueda a 100
            swdw(driver, 2, 1, "FrontsTable:ResultsDisplayed:0").click()
            swdw(driver, 2, 1, "FrontsTable:ResultsDisplayed:0").send_keys(str(100) + Keys.ENTER)
            swdw(driver, 2, 1, 'FrontsTable:ShowDetails:0')
            time.sleep(0.5)
            break

        except (StaleElementReferenceException, ElementClickInterceptedException):
            continue



def go_contract(driver):
    # Ir a buscar contratos
    
    while True:

        # Definiciones
        buscar_contract = ["JAV_MC_CAO_QRO", "Contratos", "Buscar contratos"]

        try:
            # Verificador de estar en frentes
            swdw(driver, 2, 0, "//h2[text()='Listado de Contratos']")
            break

        except TimeoutException:

            # Intentamos click en buscar frentes
            try:
                new_mainmenu(driver, buscar_contract)

            # De no poder reingresamos
            except TimeoutException:
                acceder_oracle(driver)
                new_mainmenu(driver, buscar_contract)
'''
